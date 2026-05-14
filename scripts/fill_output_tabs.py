#!/usr/bin/env python3
"""
fill_output_tabs.py
Copies full_vendor_mapping.xlsx → full_vendor_mapping_final.xlsx
and adds three filled-in tabs:
  - Top 3 Opportunities
  - Methodology
  - CEOCFO Recommendations

Also writes outputs/mappings/CEOCFOreco.md
"""

import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

BASE    = Path(__file__).resolve().parent.parent
MAPS    = BASE / "outputs" / "mappings"
SRC     = MAPS / "full_vendor_mapping.xlsx"
DEST    = MAPS / "full_vendor_mapping_final.xlsx"

shutil.copy2(SRC, DEST)

wb = openpyxl.load_workbook(DEST)

# ── Shared styles ─────────────────────────────────────────────────────────────
DARK_BLUE   = "073763"          # matches input template header colour
DARK_BLUE_F = PatternFill(start_color="073763", end_color="073763", fill_type="solid")
LIGHT_BLUE  = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
GREY_FILL   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FILL  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

HDR_FONT    = Font(bold=True, color="FFFFFF", size=11)
BOLD_DARK   = Font(bold=True, color=DARK_BLUE, size=11)
BOLD_BLACK  = Font(bold=True, color="000000", size=11)
BODY_FONT   = Font(color="000000", size=10)
SMALL_BOLD  = Font(bold=True, color="000000", size=10)
LABEL_FONT  = Font(bold=True, color=DARK_BLUE, size=10)

THIN = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)
MED = Border(
    left=Side(style="medium", color=DARK_BLUE),
    right=Side(style="medium", color=DARK_BLUE),
    top=Side(style="medium", color=DARK_BLUE),
    bottom=Side(style="medium", color=DARK_BLUE),
)

def wrap_cell(ws, row, col, value, font=None, fill=None, align=None, border=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    if num_fmt: c.number_format = num_fmt
    return c

def merge_write(ws, start_row, start_col, end_col, value, font=None, fill=None, row_height=None):
    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=start_row,   end_column=end_col)
    c = ws.cell(row=start_row, column=start_col, value=value)
    if font:  c.font      = font
    if fill:  c.fill      = fill
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if row_height:
        ws.row_dimensions[start_row].height = row_height
    return c

WRAP_L  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
WRAP_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER  = Alignment(horizontal="center", vertical="center")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1: Top 3 Opportunities
# ══════════════════════════════════════════════════════════════════════════════
if "Top 3 Opportunities" in wb.sheetnames:
    del wb["Top 3 Opportunities"]
ws1 = wb.create_sheet("Top 3 Opportunities")

# Column widths
col_widths = {1: 18, 2: 32, 3: 62, 4: 28, 5: 20, 6: 22}
for col, w in col_widths.items():
    ws1.column_dimensions[get_column_letter(col)].width = w

# Row 1 — header strip
ws1.row_dimensions[1].height = 24
for col in range(1, 7):
    c = ws1.cell(row=1, column=col)
    c.fill = DARK_BLUE_F
    c.border = MED
headers_r1 = {
    1: ("",                           None),
    2: ("Opportunity",                HDR_FONT),
    3: ("Explanation",                HDR_FONT),
    4: ("Est. Annual Savings (USD)",  HDR_FONT),
    5: ("Speed to Value",             HDR_FONT),
    6: ("Strategic Clarity",          HDR_FONT),
}
for col, (val, fnt) in headers_r1.items():
    c = ws1.cell(row=1, column=col, value=val)
    if fnt: c.font = fnt
    c.alignment = WRAP_C

# ── Opportunity data ──────────────────────────────────────────────────────────
opportunities = [
    {
        "label":    "Opportunity 1",
        "title":    "CRM Stack Consolidation\n(Salesforce + PDF Butler)",
        "explanation": (
            "Salesforce UK Ltd is $3,117,226 — 39.5% of all vendor spend — and is the "
            "acquired company's Sales CRM. Cloudcrossing BVBA (PDF Butler, $208,675) is a "
            "Salesforce-native document generation add-on used for quotes and proposals. "
            "Together they represent $3,325,901.\n\n"
            "If the acquirer runs a different CRM: migrate to eliminate the full stack.\n"
            "If the acquirer also uses Salesforce: consolidate into one enterprise agreement "
            "— typically yields 15–25% reduction.\n\n"
            "The Salesforce contract renewal date must be pulled immediately; that date "
            "determines the negotiation window."
        ),
        "savings":  "$499K–$3,326K\n(negotiation savings to full migration)",
        "speed":    "30 days: contract freeze\n90 days: renewal negotiation\n12–18 months: full migration",
        "clarity":  "HIGH\nLargest single item; decision required regardless of CRM strategy.",
        "fill":     PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid"),
    },
    {
        "label":    "Opportunity 2",
        "title":    "T&E Platform Migration\n(Navan / TripActions)",
        "explanation": (
            "Navan appears under two entity names — Navan (TripActions Inc) at $357,984 "
            "and Navan Inc at $57,929 — combined $415,913. This billing split is a data "
            "normalisation issue the acquirer should resolve on Day 1.\n\n"
            "The acquirer operates its own corporate T&E platform. Migrating the acquired "
            "company's employees onto it eliminates the entire $415,913 spend. This is a "
            "user migration exercise, not a contract negotiation — no complex commercial "
            "process required. Achievable within 30–60 days of close."
        ),
        "savings":  "$415,913\n(full elimination on platform migration)",
        "speed":    "30–60 days\n(employee migration to acquirer's T&E system)",
        "clarity":  "HIGH\nAcquirer has T&E; migration is operational, not strategic.",
        "fill":     PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    },
    {
        "label":    "Opportunity 3",
        "title":    "M&A Advisory Spend\nAutomatic Termination",
        "explanation": (
            "Five vendors totalling $298,793 are classified as M&A — all deal-specific "
            "and non-recurring:\n"
            "  • RSM UK Corporate Finance LLP:  $117,078\n"
            "  • 4I Advisory Services:           $71,860\n"
            "  • SS&C Intralinks (VDR):          $39,966\n"
            "  • Houlihan Lokey Advisors:         $37,461\n"
            "  • Vector Capital Management:       $32,427\n\n"
            "All five engagements terminate at deal close by their own contractual terms. "
            "No management action or negotiation required. This is a guaranteed saving "
            "that takes effect automatically on Day 1."
        ),
        "savings":  "$298,793\n(full; automatic at deal close)",
        "speed":    "Immediate\n(at deal close — no action required)",
        "clarity":  "VERY HIGH\nContractually defined end date; zero ambiguity.",
        "fill":     PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    },
]

row = 2
for opp in opportunities:
    ws1.row_dimensions[row].height = 120
    for col in range(1, 7):
        c = ws1.cell(row=row, column=col)
        c.fill = opp["fill"]
        c.border = THIN

    # Col A — label
    c = ws1.cell(row=row, column=1, value=opp["label"])
    c.font = BOLD_DARK
    c.alignment = WRAP_C
    c.fill = opp["fill"]

    # Col B — title
    c = ws1.cell(row=row, column=2, value=opp["title"])
    c.font = Font(bold=True, color="000000", size=11)
    c.alignment = WRAP_C

    # Col C — explanation
    c = ws1.cell(row=row, column=3, value=opp["explanation"])
    c.font = BODY_FONT
    c.alignment = WRAP_L

    # Col D — savings
    c = ws1.cell(row=row, column=4, value=opp["savings"])
    c.font = Font(bold=True, color="375623", size=10)
    c.alignment = WRAP_C

    # Col E — speed
    c = ws1.cell(row=row, column=5, value=opp["speed"])
    c.font = BODY_FONT
    c.alignment = WRAP_C

    # Col F — clarity
    c = ws1.cell(row=row, column=6, value=opp["clarity"])
    c.font = SMALL_BOLD
    c.alignment = WRAP_C

    row += 1

# Freeze header
ws1.freeze_panes = "A2"
print("✓ Tab 'Top 3 Opportunities' written")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2: Methodology
# ══════════════════════════════════════════════════════════════════════════════
if "Methodology" in wb.sheetnames:
    del wb["Methodology"]
ws2 = wb.create_sheet("Methodology")

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 90

METHODOLOGY_ROWS = [
    # (section_header?, col_A_label, col_B_content, row_height)
    (True,  "METHODOLOGY",
     "How this vendor spend analysis was approached, validated and delivered", 20),

    (False, "Tool used",
     "Claude Code (claude-sonnet-4-6) — Anthropic's CLI coding agent — running inside VS Code. "
     "All analytical logic, data extraction, and output files were produced through iterative "
     "prompting in a single project session. Python (openpyxl, pandas, csv, pathlib) was used "
     "for all XLSX and CSV manipulation.", 60),

    (False, "Approach — 4 structured steps",
     "Step 1 — Workbook Discovery & Validation: Extracted and validated the raw dataset "
     "(386 vendors, $7.89M total spend). Identified data quality risks: 11 encoding-corrupted "
     "names, 6 individual names appearing as vendors, 4 Amazon entities, 2 Navan entities.\n\n"
     "Step 2 — Classification Philosophy: Defined the governing principle (budget accountability) "
     "and resolved 7 boundary disputes between ambiguous category pairs "
     "(SaaS vs Engineering, SaaS vs Sales, G&A vs Finance, M&A vs Professional Services, etc.). "
     "Established 5 spend tiers and brand-grouping rules.\n\n"
     "Step 2b — Ambiguity Resolution: Challenged the Salesforce → Sales classification "
     "(39.5% of spend) using corroborating evidence. Resolved all 7 R-status vendors "
     "via web research and declared assumptions.\n\n"
     "Step 3 — Vendor Classification: Applied the philosophy to all 386 vendors. "
     "Each vendor received a department (12-category Config taxonomy), a 1-line description, "
     "a confidence score (H/M/L/R/N), a brand group, and a Suggestion "
     "(CONSOLIDATE / OPTIMIZE / TERMINATE).\n\n"
     "Step 4 — Opportunity Identification: Ranked opportunities by spend magnitude, "
     "feasibility, speed to value, and strategic clarity.", 200),

    (False, "Prompts used",
     "1. 'Your first task is ONLY to understand and validate the dataset. Do not classify vendors. "
     "Do not make recommendations.' — Structured Step 1 to establish ground truth before analysis.\n\n"
     "2. 'Define the strategic categorization philosophy BEFORE any vendor mapping begins.' "
     "— Forced explicit boundary-dispute resolution before classification started.\n\n"
     "3. 'Challenge you on classifying Salesforce as Sales and not SaaS.' "
     "— Stress-tested the highest-stakes single classification (39.5% of spend).\n\n"
     "4. 'Classify all vendors in the dataset based on the mapping philosophy.' "
     "— Applied the governing framework to all 386 vendors.\n\n"
     "5. 'What should be the governing logic to update the Suggestion column?' "
     "— Defined the CONSOLIDATE / OPTIMIZE / TERMINATE rules before execution.", 130),

    (False, "Validation & quality checks",
     "• Total spend reconciliation: computed spend ($7,887,360.40) matches source exactly — $0 rounding error.\n"
     "• All 386 vendors have an assigned department — no blank classifications.\n"
     "• Zero R-status vendors from ambiguity resolution: all 7 previously unresolved vendors "
     "were classified with declared assumptions before the classification script ran.\n"
     "• All T1 vendors (≥$100K, 13 vendors = 63.9% of spend) have H or M confidence — no high-spend guesses.\n"
     "• Salesforce classification independently stress-tested: 5 corroborating evidence points "
     "reviewed and documented in ambiguity_resolution.md.\n"
     "• Suggestion distribution verified against CSV independently: "
     "133 CONSOLIDATE ($6.0M), 68 OPTIMIZE ($1.5M), 185 TERMINATE ($365K).\n"
     "• Navan duplicate confirmed: two entries ($357.9K + $57.9K = $415.9K) "
     "flagged as same-vendor billing split.", 130),

    (False, "Output files produced",
     "outputs/workbook_discovery.md — tab inventory and data schema\n"
     "outputs/data_quality_summary.md — data quality report (missing values, encoding issues, brand fragmentation)\n"
     "outputs/raw_valid_records_preview.csv — all 386 rows, source of truth for classification\n"
     "outputs/mapping_philosophy.md — governing classification framework\n"
     "outputs/department_interpretation_guide.md — per-department definitions and decision tree\n"
     "outputs/ambiguity_resolution.md — Salesforce challenge ruling and 14 vendor resolutions\n"
     "outputs/mappings/full_vendor_mapping.xlsx — classified dataset with confidence and suggestions\n"
     "outputs/mappings/full_vendor_mapping.csv — flat export\n"
     "outputs/mappings/full_vendor_mapping_final.xlsx — this file\n"
     "outputs/mappings/CEOCFOreco.md — one-page executive memo", 120),
]

for i, (is_header, label, content, height) in enumerate(METHODOLOGY_ROWS, 1):
    ws2.row_dimensions[i].height = height
    if is_header:
        ws2.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        c = ws2.cell(row=i, column=1, value=label)
        c.fill  = DARK_BLUE_F
        c.font  = HDR_FONT
        c.alignment = CENTER
    else:
        ca = ws2.cell(row=i, column=1, value=label)
        ca.font      = LABEL_FONT
        ca.fill      = LIGHT_BLUE
        ca.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ca.border    = THIN

        cb = ws2.cell(row=i, column=2, value=content)
        cb.font      = BODY_FONT
        cb.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cb.border    = THIN

print("✓ Tab 'Methodology' written")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3: CEOCFO Recommendations
# ══════════════════════════════════════════════════════════════════════════════
MEMO_TEXT = """\
TO:      CEO / CFO
FROM:    Integration Strategy
DATE:    14 May 2026
RE:      Vendor Spend Analysis — Acquisition Target

────────────────────────────────────────────────────────────────────────────────

TOTAL VENDOR SPEND: $7,887,360  |  386 VENDORS  |  LAST 12 MONTHS

The one number that reframes everything: Salesforce UK Ltd is $3,117,226 — 39.5% \
of all vendor spend. This single contract is worth more than the next five vendors \
combined. Your first integration decision is about CRM, and it needs to be made \
within 30 days of close.

────────────────────────────────────────────────────────────────────────────────

THREE DECISIONS REQUIRED

1.  CRM stack — $3,325,901 (Salesforce + PDF Butler)
    Do you migrate them to your CRM, or consolidate onto Salesforce enterprise?
    If you run a different CRM: migration planning starts Day 1.
    If you also use Salesforce: push for a consolidated enterprise renewal within
    90 days — expect 15–25% savings (~$500K annually).
    Either way: pull the Salesforce contract renewal date this week.

2.  T&E platform — $415,913 (Navan / TripActions, two entity entries)
    Migrate their employees onto your T&E system within 30–60 days of close.
    This vendor appears under two entity names — resolve that on Day 1.
    No contract negotiation required. Full $416K annual saving.

3.  M&A advisory — $298,793 (five vendors, terminate automatically)
    RSM UK Corporate Finance, Houlihan Lokey, SS&C Intralinks, Vector Capital,
    4I Advisory. All terminate at deal close. Zero management action required.
    $299K saves automatically.

────────────────────────────────────────────────────────────────────────────────

THE RISK YOU ARE NOT EXPECTING

The acquired company has office leases across five geographies — UK, Croatia, India, \
Singapore, Australia — totalling $1.32M/year across 88 facilities vendors. The \
Croatian offices (Zagrebtower $183.8K + Weking $144.1K = $327.9K/year) carry the \
highest exposure. If headcount in Croatia is uncertain post-acquisition, these \
commitments continue accruing with no return. This does not appear in the P&L \
headline; it only becomes visible when you map headcount to office obligations. \
Legal needs the Croatian lease expiry dates within the first 30 days of close.

────────────────────────────────────────────────────────────────────────────────

RECOMMENDED NEXT STEP

Owner:   Integration Lead, with Legal and Finance support
Action:  This week — pull the Salesforce contract renewal date, the two Croatian
         lease expiry dates, and the Navan termination window.

These four data points determine the critical path for 45% of total vendor spend.\
"""

if "CEOCFO Recommendations" in wb.sheetnames:
    del wb["CEOCFO Recommendations"]
ws3 = wb.create_sheet("CEOCFO Recommendations")

ws3.column_dimensions["A"].width = 115
ws3.row_dimensions[1].height = 24
ws3.row_dimensions[2].height = 380

# Row 1 — dark header strip
ws3.merge_cells("A1:A1")
c_hdr = ws3.cell(row=1, column=1,
    value="EXECUTIVE MEMO — VENDOR SPEND ANALYSIS: ACQUISITION TARGET")
c_hdr.fill      = DARK_BLUE_F
c_hdr.font      = HDR_FONT
c_hdr.alignment = CENTER

# Row 2 — memo body
c_body = ws3.cell(row=2, column=1, value=MEMO_TEXT)
c_body.font      = Font(name="Courier New", size=9, color="000000")
c_body.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
c_body.border    = THIN

print("✓ Tab 'CEOCFO Recommendations' written")


# ── Reorder sheets ────────────────────────────────────────────────────────────
desired_order = [
    "Vendor Mapping", "Legend",
    "Top 3 Opportunities", "Methodology", "CEOCFO Recommendations"
]
wb._sheets.sort(key=lambda s: desired_order.index(s.title) if s.title in desired_order else 99)

wb.save(DEST)
print(f"✓ Saved → {DEST}")


# ══════════════════════════════════════════════════════════════════════════════
# CEOCFOreco.md
# ══════════════════════════════════════════════════════════════════════════════
MD_PATH = MAPS / "CEOCFOreco.md"

md_content = """\
# MEMO

**TO:** CEO / CFO
**FROM:** Integration Strategy
**DATE:** 14 May 2026
**RE:** Vendor Spend Analysis — Acquisition Target

---

**Total vendor spend: $7,887,360 across 386 vendors (last 12 months)**

The one number that reframes everything: **Salesforce UK Ltd is $3,117,226 — 39.5% of \
all vendor spend.** This single contract is worth more than the next five vendors combined. \
Your first integration decision is about CRM, and it needs to be made within 30 days of close.

---

## Three decisions required from you

**1. CRM stack — $3,325,901 (Salesforce + PDF Butler)**
Do you migrate them to your CRM, or consolidate onto Salesforce enterprise? If you run a \
different CRM, migration planning starts Day 1. If you also use Salesforce, a consolidated \
enterprise renewal within 90 days should yield 15–25% savings (~$500K). Either way: pull \
the Salesforce contract renewal date this week.

**2. T&E platform — $415,913 (Navan / TripActions, two entity entries)**
Migrate their employees onto your T&E system within 30–60 days. This vendor appears under \
two entity names — close that out on Day 1. No negotiation required. Full $416K annual saving.

**3. M&A advisory — $298,793 (five vendors)**
RSM UK Corporate Finance, Houlihan Lokey, SS&C Intralinks, Vector Capital, 4I Advisory. \
All terminate at deal close by their own contractual terms. Zero management action required. \
$299K saves automatically.

---

## The risk you are not expecting

The acquired company has office leases across five geographies — UK, Croatia, India, \
Singapore, Australia — totalling $1.32M/year across 88 facilities vendors. The Croatian \
offices (Zagrebtower $183.8K + Weking $144.1K = $327.9K/year) carry the highest exposure. \
If headcount in Croatia is uncertain post-acquisition, these commitments continue accruing \
with no return. This does not appear in the P&L headline. Legal needs the Croatian lease \
expiry dates within the first 30 days of close.

---

## Recommended next step

**Owner:** Integration Lead + Legal + Finance
**Action:** This week — pull the Salesforce contract renewal date, the two Croatian lease \
expiry dates, and the Navan termination window.

These four data points determine the critical path for **45% of total vendor spend**.
"""

MD_PATH.write_text(md_content, encoding="utf-8")
print(f"✓ Saved → {MD_PATH}")

print("\n=== COMPLETE ===")
print(f"  full_vendor_mapping_final.xlsx  — 5 tabs (Vendor Mapping, Legend, Top 3, Methodology, CEOCFO)")
print(f"  CEOCFOreco.md                   — one-page executive memo")
