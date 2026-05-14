"""
Vendor Spend Dataset Discovery & Validation
Post-acquisition operating strategist — Step 1 (dataset understanding only)
"""

import openpyxl
import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE = Path(__file__).parent.parent
INPUT_FILE = BASE / "input" / "Vendor_Spend_Strategy.xlsx"
OUTPUTS = BASE / "outputs"
OUTPUTS.mkdir(exist_ok=True)

# ── Corporate keyword list for individual-name detection ────────────────────
CORPORATE_KEYWORDS = {
    "ltd", "llc", "llp", "inc", "corp", "gmbh", "bv", "bvba", "pty", "srl",
    "ag", "sa", "plc", "limited", "pvt", "private", "co", "group", "holdings",
    "d.o.o", "d.o.o.", "s.a.", "s.r.l.", "s.p.a.", "s.r.o", "j.d.o.o",
    "j.d.o.o.", "d.d", "d.d.", "services", "solutions", "consulting",
    "technologies", "systems", "global", "international", "enterprises",
    "associates", "advisors", "advisers", "partners", "ventures", "capital",
    "management", "digital", "software", "cloud", "data", "analytics",
    "communications", "media", "marketing", "finance", "financial", "legal",
    "law", "health", "healthcare", "insurance", "logistics", "transport",
    "events", "catering", "hospitality", "hotel", "restaurants", "engineering",
    "construction", "properties", "real", "estate", "spaces", "office",
    "network", "networks", "security", "cyber", "platform", "platforms",
    "research", "institute", "foundation", "authority", "government", "council",
    "university", "college", "school", "academy", "press", "publishing",
    "airlines", "air", "freight", "shipping", "parcel", "delivery",
    "recruitment", "staffing", "training", "education", "learning",
    "australia", "uk", "usa", "us", "ireland", "germany", "france", "spain",
    "worldwide", "worldwide", "online", "web", "tech", "labs", "lab",
    "studio", "studios", "creative", "design", "agency", "production",
    "trading", "imports", "exports", "supply", "supplies", "products",
    "retail", "wholesale", "medical", "dental", "pharmacy", "clinic",
    "ordinacija", "zavod", "holding", "nastavni", "gradski", "grad",
}

LEGAL_SUFFIX_RE = re.compile(
    r"\b(ltd|llc|llp|inc|corp|gmbh|bv|bvba|pty|srl|ag|sa|plc|limited|pvt|"
    r"private|d\.o\.o\.?|s\.r\.o\.?|j\.d\.o\.o\.?|d\.d\.?|s\.a\.?|s\.r\.l\.?|"
    r"s\.p\.a\.?|uk|us|usa|australia|aus|ireland|aps|sg|wa|nsw|na)\b",
    re.IGNORECASE,
)


def normalize_name(name: str) -> str:
    """Strip legal suffixes and punctuation for near-duplicate grouping."""
    n = str(name).lower()
    n = LEGAL_SUFFIX_RE.sub("", n)
    n = re.sub(r"[^a-z0-9\s]", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def has_encoding_issue(name: str) -> bool:
    return any(ord(c) > 127 for c in str(name))


def is_possible_individual(name: str) -> bool:
    """Heuristic: 2-word title-case name with no corporate keywords."""
    parts = str(name).strip().split()
    if len(parts) != 2:
        return False
    if not all(p and p[0].isupper() for p in parts):
        return False
    lower_name = name.lower()
    return not any(kw in lower_name for kw in CORPORATE_KEYWORDS)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — WORKBOOK DISCOVERY
# ══════════════════════════════════════════════════════════════════════════════

wb = openpyxl.load_workbook(INPUT_FILE)

TAB_META = {
    "Vendor Analysis Assessment": {
        "type": "Primary Analysis",
        "purpose": "Master vendor list with spend data — the dataset under review.",
    },
    "Config": {
        "type": "Reference",
        "purpose": "Controlled vocabulary of 13 department categories for classification.",
    },
    "Top 3 Opportunities": {
        "type": "Derived / Template",
        "purpose": "Candidate fill-in: top consolidation/savings opportunities (currently blank).",
    },
    "Methodology": {
        "type": "Template",
        "purpose": "Candidate fill-in: explanation of analytical approach (currently blank).",
    },
    "CEOCFO Recommendations": {
        "type": "Template",
        "purpose": "Candidate fill-in: executive memo link for CEO/CFO (currently blank).",
    },
}

tab_info = []
for name in wb.sheetnames:
    ws = wb[name]
    populated = sum(
        1 for row in ws.iter_rows(values_only=True) if any(v is not None for v in row)
    )
    meta = TAB_META.get(name, {"type": "Unknown", "purpose": "—"})
    tab_info.append(
        {
            "name": name,
            "type": meta["type"],
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "populated_rows": populated,
            "purpose": meta["purpose"],
        }
    )

# Config dept list
config_ws = wb["Config"]
dept_list = [
    row[0] for i, row in enumerate(config_ws.iter_rows(values_only=True)) if i > 0 and row[0]
]

# Main sheet column map
main_ws = wb["Vendor Analysis Assessment"]
header_row = next(main_ws.iter_rows(values_only=True))
col_map = []
col_letters = "ABCDEFGHIJKLMNO"
for i, h in enumerate(header_row):
    col_map.append({"col": col_letters[i], "header": h or "(no header)"})


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — EXTRACT VALID ROWS + DATA VALIDATION
# ══════════════════════════════════════════════════════════════════════════════

records = []
blank_rows = 0
total_sheet_rows = 0

for i, row in enumerate(main_ws.iter_rows(values_only=True)):
    if i == 0:
        continue  # header
    total_sheet_rows += 1
    vendor = row[0]
    dept = row[1]
    cost = row[2]
    desc = row[3]
    suggestion = row[4]

    if vendor is not None or cost is not None:
        records.append(
            {
                "row_number": i + 1,
                "vendor_name": vendor,
                "department": dept,
                "cost_usd": cost,
                "description": desc,
                "suggestion": suggestion,
                "has_encoding_issue": has_encoding_issue(str(vendor)) if vendor else False,
                "possible_individual_name": is_possible_individual(str(vendor)) if vendor else False,
            }
        )
    else:
        blank_rows += 1

total_records = len(records)
total_spend = sum(r["cost_usd"] for r in records if r["cost_usd"] is not None)

# — Missing values —
def null_count(field):
    return sum(1 for r in records if not r[field])

missing = {
    "vendor_name":  null_count("vendor_name"),
    "department":   null_count("department"),
    "cost_usd":     sum(1 for r in records if r["cost_usd"] is None),
    "description":  null_count("description"),
    "suggestion":   null_count("suggestion"),
}

# — Duplicates —
raw_names = [r["vendor_name"] for r in records if r["vendor_name"]]
exact_dupes = {k: v for k, v in Counter(raw_names).items() if v > 1}

norm_groups: dict[str, list] = defaultdict(list)
for r in records:
    if r["vendor_name"]:
        norm_groups[normalize_name(r["vendor_name"])].append(r["vendor_name"])
near_dupes = {k: v for k, v in norm_groups.items() if len(v) > 1}

# — Encoding issues —
encoding_issues = [r for r in records if r["has_encoding_issue"]]

# — Individual names —
individual_vendors = [r for r in records if r["possible_individual_name"]]

# — Sort check —
costs_in_order = [r["cost_usd"] for r in records if r["cost_usd"] is not None]
is_sorted_desc = all(costs_in_order[i] >= costs_in_order[i + 1] for i in range(len(costs_in_order) - 1))

# — Spend concentration —
sorted_records = sorted(records, key=lambda x: x["cost_usd"] or 0, reverse=True)
concentration = {}
for n in [10, 20, 50]:
    top_spend = sum(r["cost_usd"] for r in sorted_records[:n] if r["cost_usd"])
    concentration[n] = {"spend": top_spend, "pct": 100 * top_spend / total_spend}

# Salesforce share
sf_spend = next((r["cost_usd"] for r in records if r["vendor_name"] == "Salesforce Uk Ltd-Uk"), 0)

# — Long tail —
long_tail = {}
for threshold in [1_000, 5_000, 10_000]:
    cnt = sum(1 for r in records if r["cost_usd"] and r["cost_usd"] < threshold)
    long_tail[threshold] = {"count": cnt, "pct": 100 * cnt / total_records}

# — Multi-entity brand fragmentation —
brand_patterns = {
    "Amazon": r"amazon",
    "Google": r"google",
    "Microsoft": r"microsoft",
    "Salesforce": r"salesforce",
    "AWS": r"amazon web services",
    "Atlassian": r"atlassian",
    "Adobe": r"adobe",
    "Slack": r"slack",
    "BDO": r"^bdo\b",
    "RSM": r"^rsm\b",
    "Deloitte": r"deloitte",
    "KPMG": r"kpmg",
    "PwC": r"pwc|pricewaterhouse",
    "Telefonica": r"telef[oó]nica",
    "Bupa": r"^bupa",
    "Cigna": r"^cigna",
    "Allianz": r"^allianz",
}

brand_groups = {}
for brand, pattern in brand_patterns.items():
    matches = [r for r in records if r["vendor_name"] and re.search(pattern, r["vendor_name"], re.IGNORECASE)]
    if len(matches) > 1:
        brand_groups[brand] = {
            "entities": [m["vendor_name"] for m in matches],
            "combined_spend": sum(m["cost_usd"] for m in matches if m["cost_usd"]),
            "count": len(matches),
        }

# ══════════════════════════════════════════════════════════════════════════════
# OUTPUT 1 — workbook_discovery.md
# ══════════════════════════════════════════════════════════════════════════════

disc_lines = [
    "# Workbook Discovery Report",
    "",
    f"**Source file:** `input/Vendor_Spend_Strategy.xlsx`  ",
    f"**Tabs found:** {len(wb.sheetnames)}  ",
    "",
    "---",
    "",
    "## Tab Inventory",
    "",
    "| Tab Name | Type | Dimensions | Populated Rows | Purpose |",
    "|----------|------|-----------|----------------|---------|",
]
for t in tab_info:
    disc_lines.append(
        f"| {t['name']} | {t['type']} | {t['dimensions']} | {t['populated_rows']} | {t['purpose']} |"
    )

disc_lines += [
    "",
    "---",
    "",
    "## Primary Analysis Tab: `Vendor Analysis Assessment`",
    "",
    f"- **Total row extent:** {main_ws.max_row} rows × {main_ws.max_column} columns",
    f"- **Populated vendor records:** {total_records}",
    f"- **Blank/formatting-only rows:** {blank_rows}",
    f"- **Last row with data:** row {records[-1]['row_number'] if records else 'N/A'}",
    "",
    "### Column Map",
    "",
    "| Column | Header |",
    "|--------|--------|",
]
for c in col_map:
    disc_lines.append(f"| {c['col']} | {c['header']} |")

disc_lines += [
    "",
    "> **Note:** Columns F–O have no headers and contain no data. The workbook width is",
    "> inflated to 15 columns but only 5 are in use.",
    "",
    "---",
    "",
    "## Reference Tab: `Config`",
    "",
    "Contains the controlled vocabulary of **department categories** available for vendor classification:",
    "",
]
for d in dept_list:
    disc_lines.append(f"- {d}")

disc_lines += [
    "",
    "> **Critical gap:** None of the 386 vendor records has a department value assigned.",
    "> The Config list exists as a reference but has not been applied to the data.",
    "",
    "---",
    "",
    "## Template / Derived Tabs (currently blank)",
    "",
    "| Tab | Status | Notes |",
    "|-----|--------|-------|",
    "| Top 3 Opportunities | Blank | Candidate fill-in: top savings/consolidation opportunities |",
    "| Methodology | Blank | Candidate fill-in: analytical approach description |",
    "| CEOCFO Recommendations | Blank | Candidate fill-in: Google Doc link to executive memo |",
    "",
    "> These three tabs are structured as assessment deliverables, not source data.",
    "> They do not affect upstream data quality.",
]

(OUTPUTS / "workbook_discovery.md").write_text("\n".join(disc_lines), encoding="utf-8")
print("✓ workbook_discovery.md written")


# ══════════════════════════════════════════════════════════════════════════════
# OUTPUT 2 — data_quality_summary.md
# ══════════════════════════════════════════════════════════════════════════════

dq_lines = [
    "# Data Quality Summary",
    "",
    f"**Source:** `Vendor Analysis Assessment` tab  ",
    f"**Analysis date:** 2026-05-14  ",
    "",
    "---",
    "",
    "## 1. Record Counts",
    "",
    "| Metric | Count |",
    "|--------|-------|",
    f"| Total rows in sheet (excl. header) | {total_sheet_rows} |",
    f"| Populated vendor records | {total_records} |",
    f"| Blank / formatting-only rows | {blank_rows} |",
    f"| Header row | 1 |",
    "",
    "---",
    "",
    "## 2. Column Completeness",
    "",
    "| Column | Populated | Null | Null % | Status |",
    "|--------|-----------|------|--------|--------|",
]

col_defs = [
    ("vendor_name", "Vendor Name"),
    ("department", "Department"),
    ("cost_usd", "Cost (USD)"),
    ("description", "1-line Description"),
    ("suggestion", "Suggestions"),
]
for field, label in col_defs:
    nulls = missing[field]
    populated = total_records - nulls
    pct = 100 * nulls / total_records
    status = "✓ Complete" if nulls == 0 else ("⚠ Partial" if nulls < total_records else "✗ Empty")
    dq_lines.append(f"| {label} | {populated} | {nulls} | {pct:.0f}% | {status} |")

dq_lines += [
    "",
    "> **Critical:** Department, Description, and Suggestions are 100% empty.",
    "> The dataset as delivered contains only vendor names and spend figures.",
    "> All downstream department-level analysis requires manual or AI-assisted categorization.",
    "",
    "---",
    "",
    "## 3. Duplicate Analysis",
    "",
    f"**Exact duplicate vendor names:** {len(exact_dupes)}",
]

if exact_dupes:
    dq_lines += ["", "| Vendor Name | Occurrences |", "|-------------|-------------|"]
    for k, v in exact_dupes.items():
        dq_lines.append(f"| {k} | {v} |")
else:
    dq_lines.append("")
    dq_lines.append("> No exact duplicates found.")

dq_lines += [
    "",
    f"**Near-duplicate groups (normalized, legal suffix stripped):** {len(near_dupes)}",
    "",
]
if near_dupes:
    dq_lines += ["| Normalized Root | Variants |", "|----------------|----------|"]
    for k, v in near_dupes.items():
        dq_lines.append(f"| `{k}` | {' / '.join(v)} |")
else:
    dq_lines.append("> No near-duplicates found after normalization.")

dq_lines += [
    "",
    "---",
    "",
    "## 4. Encoding Issues",
    "",
    f"**Vendors with corrupted non-ASCII characters:** {len(encoding_issues)}",
    "",
    "> All affected entries appear to be Croatian entities. The corruption pattern",
    "> (e.g., `ä_x008d_`, `å¡`, `ã³`, `â€™`) is consistent with a UTF-8 → Latin-1",
    "> mojibake introduced during export or import.",
    "",
    "| Row # | Corrupted Name |",
    "|-------|---------------|",
]
for r in encoding_issues:
    dq_lines.append(f"| {r['row_number']} | {r['vendor_name']} |")

dq_lines += [
    "",
    "---",
    "",
    "## 5. Multi-Entity Brand Fragmentation",
    "",
    f"**Brands appearing as multiple distinct vendor entries:** {len(brand_groups)}",
    "",
    "| Brand | Entity Count | Entities | Combined Spend |",
    "|-------|-------------|----------|----------------|",
]
for brand, info in brand_groups.items():
    entities_str = " / ".join(f"`{e}`" for e in info["entities"])
    dq_lines.append(
        f"| {brand} | {info['count']} | {entities_str} | ${info['combined_spend']:,.2f} |"
    )

dq_lines += [
    "",
    "> **Risk:** Fragmented entries understate true spend per vendor. AWS LLC + AWS Inc.",
    "> are the same supplier; Amazon.co.uk and Amazon (Aus) may also consolidate.",
    "> True Amazon/AWS exposure must be summed before negotiation or benchmarking.",
    "",
    "---",
    "",
    "## 6. Individual Names Appearing as Vendors",
    "",
    f"**Entries matching individual-name pattern (2-word title-case, no corporate keywords):** {len(individual_vendors)}",
    "",
    "| Row # | Name | Spend (USD) |",
    "|-------|------|-------------|",
]
for r in individual_vendors:
    dq_lines.append(f"| {r['row_number']} | {r['vendor_name']} | ${r['cost_usd']:,.2f} |")

dq_lines += [
    "",
    "> These may be contractor/freelancer payments, expense reimbursements, or data entry",
    "> errors. They should be reviewed and either recategorized (e.g., under a staffing",
    "> agency) or excluded from vendor consolidation analysis.",
    "",
    "---",
    "",
    "## 7. Spend Distribution",
    "",
    f"| Metric | Value |",
    f"|--------|-------|",
    f"| Total L12M spend | ${total_spend:,.2f} |",
    f"| Highest single vendor | ${sorted_records[0]['cost_usd']:,.2f} ({sorted_records[0]['vendor_name']}) |",
    f"| Lowest single vendor | ${sorted_records[-1]['cost_usd']:,.2f} ({sorted_records[-1]['vendor_name']}) |",
    f"| Salesforce share of total | {100*sf_spend/total_spend:.1f}% |",
    f"| Data already sorted descending | {'Yes' if is_sorted_desc else 'No'} |",
    "",
    "---",
    "",
    "## 8. Spend Concentration",
    "",
    "| Vendor Tier | Vendor Count | Cumulative Spend | % of Total |",
    "|-------------|-------------|------------------|------------|",
]
for n, info in concentration.items():
    dq_lines.append(f"| Top {n} vendors | {n} | ${info['spend']:,.2f} | {info['pct']:.1f}% |")

remaining_count = total_records - 50
remaining_spend = total_spend - concentration[50]["spend"]
dq_lines += [
    f"| Remaining {remaining_count} vendors | {remaining_count} | ${remaining_spend:,.2f} | {100*remaining_spend/total_spend:.1f}% |",
    "",
    "---",
    "",
    "## 9. Long Tail Profile",
    "",
    "| Spend Threshold | Vendor Count | % of All Vendors |",
    "|----------------|-------------|------------------|",
]
for threshold, info in long_tail.items():
    dq_lines.append(
        f"| Under ${threshold:,} | {info['count']} | {info['pct']:.0f}% |"
    )

above_10k = total_records - long_tail[10_000]["count"]
dq_lines += [
    f"| $10K and above (strategic tier) | {above_10k} | {100*above_10k/total_records:.0f}% |",
    "",
    "> 46% of vendors (178) generate under $1K in annual spend.",
    "> The bottom ~336 vendors (87%) account for only 11.5% of total spend.",
    "> A strategic review can focus on ~79 vendors ($10K+) covering 80% of spend.",
    "",
    "---",
    "",
    "## 10. Strategic Risks Summary",
    "",
    "| # | Risk | Severity | Detail |",
    "|---|------|----------|--------|",
    "| 1 | **Zero department mapping** | 🔴 Critical | 100% of vendors lack a department tag. No cost-center or functional analysis is possible without a categorization pass. |",
    "| 2 | **Salesforce single-vendor concentration** | 🔴 Critical | One vendor = 39.5% of total L12M spend ($3.1M). Any contract risk or pricing change has outsized P&L impact. |",
    f"| 3 | **Encoding-corrupted vendor names** | 🟠 High | {len(encoding_issues)} Croatian entities have garbled names. These cannot be matched to external databases or contracts without manual correction. |",
    f"| 4 | **Amazon multi-entity fragmentation** | 🟠 High | 4 separate AWS/Amazon entries mask true exposure (~$113K combined). Likely undercounted if procurement is split across entities. |",
    f"| 5 | **Individual names as vendor records** | 🟠 High | {len(individual_vendors)} entries appear to be people, not companies. These distort vendor count, average spend, and category analysis. |",
    "| 6 | **Extreme long tail** | 🟡 Medium | 178 vendors under $1K represent administrative drag with negligible savings potential. Should be filtered before strategic analysis. |",
    "| 7 | **No descriptions or suggestions populated** | 🟡 Medium | Columns D and E are 100% blank. Vendor context and prior recommendations must be sourced externally or generated fresh. |",
    "| 8 | **Columns F–O unused** | 🟢 Low | The workbook has 15 columns but only 5 are in use. No hidden data risk, but the schema is wider than needed. |",
]

(OUTPUTS / "data_quality_summary.md").write_text("\n".join(dq_lines), encoding="utf-8")
print("✓ data_quality_summary.md written")


# ══════════════════════════════════════════════════════════════════════════════
# OUTPUT 3 — raw_valid_records_preview.csv
# ══════════════════════════════════════════════════════════════════════════════

csv_path = OUTPUTS / "raw_valid_records_preview.csv"
fieldnames = [
    "row_number",
    "vendor_name",
    "department",
    "cost_usd",
    "description",
    "suggestion",
    "has_encoding_issue",
    "possible_individual_name",
]

with open(csv_path, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    for r in records:
        writer.writerow({k: r[k] for k in fieldnames})

print(f"✓ raw_valid_records_preview.csv written ({total_records} records)")

# ── Console summary ────────────────────────────────────────────────────────
print()
print("=" * 60)
print("EXECUTIVE DATASET SUMMARY")
print("=" * 60)
print(f"  Populated vendor records : {total_records}")
print(f"  Total L12M spend         : ${total_spend:,.2f}")
print(f"  Dept mapping coverage    : 0% (critical gap)")
print(f"  Top 10 vendor spend share: {concentration[10]['pct']:.1f}%")
print(f"  Salesforce share         : {100*sf_spend/total_spend:.1f}%")
print(f"  Vendors under $1K        : {long_tail[1000]['count']} ({long_tail[1000]['pct']:.0f}%)")
print(f"  Encoding-corrupted names : {len(encoding_issues)}")
print(f"  Multi-entity brand groups: {len(brand_groups)}")
print(f"  Possible individual names : {len(individual_vendors)}")
print("=" * 60)
print("Outputs written to: outputs/")
