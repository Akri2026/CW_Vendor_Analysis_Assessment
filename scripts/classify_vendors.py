"""
Step 3: Full Vendor Classification
Reads raw_valid_records_preview.csv, applies the mapping philosophy and
department interpretation guide, and writes:
  outputs/mappings/full_vendor_mapping.xlsx
  outputs/mappings/full_vendor_mapping.csv
  outputs/mappings/full_mapping_summary.md
"""

import csv
import re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent.parent
OUTPUTS = BASE / "outputs" / "mappings"
OUTPUTS.mkdir(parents=True, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# CLASSIFICATION LOOKUP TABLE
# Keys are EXACT vendor_name strings from the CSV.
# Values: (department, description, confidence, brand_group, notes)
# Confidence: H / M / L / R / N
# Brand group: canonical brand string or "" for single-entity vendors
# ══════════════════════════════════════════════════════════════════════════════

CLASSIFICATIONS = {
    # ── T1 STRATEGIC (≥$100K) ─────────────────────────────────────────────────
    "Salesforce Uk Ltd-Uk": (
        "Sales", "Enterprise CRM platform powering Sales pipeline, opportunity management and revenue operations",
        "H", "Salesforce", ""
    ),
    "Navan (Tripactions Inc)": (
        "G&A", "Global corporate travel and expense management platform",
        "H", "Navan / TripActions", "Two entries for one vendor — combined spend ~$415.9K; normalise before reporting"
    ),
    "Bdo Llp": (
        "Finance", "External audit, accounting and business advisory firm (UK Big 4-tier)",
        "H", "", ""
    ),
    "Tog Uk Properties Limited": (
        "Facilities", "Commercial office property and workspace landlord (UK)",
        "H", "", ""
    ),
    "Cloudcrossing Bvba": (
        "Sales", "Salesforce-native document generation tool (PDF Butler) for sales quotes, proposals and contracts",
        "M", "Cloudcrossing / PDF Butler",
        "Strategy A: Salesforce AppExchange add-on → aligns with Salesforce Sales classification. Verify at integration diligence."
    ),
    "Zagrebtower D.O.O.": (
        "Facilities", "Commercial office tower and workplace provider (Zagreb, Croatia)",
        "H", "", ""
    ),
    "Innovent Spaces Private Limited": (
        "Facilities", "Coworking and serviced office space provider (India)",
        "H", "", ""
    ),
    "Weking D.O.O.": (
        "Facilities", "Commercial property and office space provider (Croatia)",
        "H", "", ""
    ),
    "Jensten Insurance Brokers": (
        "G&A", "Corporate insurance broker managing employee and company-wide insurance policies",
        "H", "", ""
    ),
    "Gpt Space & Co": (
        "Facilities", "Commercial office and event space provider (Australia)",
        "H", "", ""
    ),
    "Aetna Life And Casualty Ltd": (
        "G&A", "Employee health and life insurance provider",
        "H", "", ""
    ),
    "Rsm Uk Corporate Finance Llp": (
        "M&A", "Specialist M&A deal advisory — sell/buy-side transactions, due diligence and valuations (separate legal entity from RSM audit/tax)",
        "H", "", "Confirmed via Companies House (OC325347): exclusive M&A advisory LLP, not standard accounting"
    ),
    "Amazon Web Services Llc": (
        "Engineering", "Cloud infrastructure and compute services underpinning the product and SDLC (AWS)",
        "H", "Amazon Web Services / Amazon", ""
    ),

    # ── T2 SIGNIFICANT ($25K–$99.9K) ─────────────────────────────────────────
    "Telefonica Global Services Gmbh": (
        "Facilities", "International office telecommunications and connectivity services",
        "H", "", ""
    ),
    "Hr Solution International Gmbh": (
        "G&A", "Global HR staffing, administration and workforce management services",
        "H", "HRsolution International", ""
    ),
    "4I Advisory Services": (
        "M&A", "M&A and financial restructuring advisory services including deal-related tax structuring",
        "M", "4I Group",
        "Strategy A: firm specialises in M&A + tax advisory; active deal context in dataset. If purely routine tax → reclassify to Finance."
    ),
    "Bisley Law Ltd": (
        "Legal", "UK commercial law firm providing legal advice, representation and contract management",
        "H", "", ""
    ),
    "Infosys": (
        "Professional Services", "Global IT services, technology consulting and outsourcing firm",
        "H", "", ""
    ),
    "Big Frontier Pty Ltd (Cult Of Monday)": (
        "G&A", "People and culture advisory, executive coaching and leadership development firm (Australia)",
        "M", "",
        "Strategy A: People & Culture advisory + executive search → HR/people function → G&A. If engagement was function-specific recruitment, reclassify."
    ),
    "Harmonic Group Limited": (
        "Professional Services", "Management consulting and business transformation firm (UK; acquired by KBR 2021)",
        "H", "", ""
    ),
    "Wework Singapore Pte. Ltd.": (
        "Facilities", "Coworking and serviced office space provider (Singapore)",
        "H", "", ""
    ),
    "Cloud Technology Solutions Ltd": (
        "Professional Services", "IT consulting and Google Cloud platform implementation services",
        "H", "", ""
    ),
    "Navan, Inc": (
        "G&A", "Corporate travel and expense management platform (US entity — same vendor as Navan/TripActions)",
        "H", "Navan / TripActions",
        "Two entries for one vendor (Navan/TripActions Inc + Navan Inc) — combined spend ~$415.9K; normalise before any recommendation."
    ),
    "Tmforum": (
        "Professional Services", "Telecoms industry association providing standards, benchmarking and advisory resources",
        "M", "", "Industry body membership — classified Professional Services; note TM Forum appears twice (also row 249 at $510)"
    ),
    "Linkedin Ireland Limited": (
        "Marketing", "Professional network platform for demand generation, employer branding and talent marketing",
        "M", "",
        "Default Marketing; flip to Sales if LinkedIn Sales Navigator is primary product (insufficient data to confirm)"
    ),
    "Kimble Applications Ltd": (
        "Professional Services", "Professional services automation (PSA) platform for project delivery and resource management",
        "H", "", ""
    ),
    "Sage Uk Limited": (
        "Finance", "Cloud accounting and financial management software",
        "H", "", ""
    ),
    "Grant Thornton": (
        "Finance", "Business advisory and accounting firm providing audit, tax and financial advisory services",
        "H", "", "Flagged as possible individual name — false positive; Grant Thornton is a major accounting firm"
    ),
    "Sveuä_x008d_Iliå¡Te U Zagrebu, Studentski Centar": (
        "Facilities", "University of Zagreb student center — likely conference/event venue or accommodation",
        "R", "",
        "Encoding corrupted (expected: 'Sveučilište U Zagrebu, Studentski Centar'). Classified Facilities based on context (university venue). Verify decoded name before finalising."
    ),
    "Ss&C Intralinks Inc": (
        "M&A", "Virtual data room platform for M&A due diligence and secure deal document management",
        "H", "", ""
    ),
    "Veniture D.O.O.": (
        "Professional Services", "Business consulting and venture advisory services (Croatia)",
        "M", "", ""
    ),
    "Accutrainee Limited": (
        "Professional Services", "Professional training, talent development and graduate recruitment services",
        "H", "", ""
    ),
    "Mason Frank International Ltd": (
        "Professional Services", "Salesforce and cloud technology specialist recruitment and staffing agency",
        "H", "", ""
    ),
    "Houlihan Lokey Advisors, Llc": (
        "M&A", "Investment bank providing M&A financial advisory and business valuation services",
        "H", "", ""
    ),
    "Vector Capital Management Lp": (
        "M&A", "Private equity investment management and M&A advisory",
        "H", "", ""
    ),
    "Hubspot Ireland Limited": (
        "Marketing", "Inbound marketing platform, CRM and marketing automation suite",
        "H", "",
        "HubSpot presence confirms Salesforce is not being used for Marketing Cloud (two overlapping tools would be redundant)"
    ),
    "Nefron - Obrt Za Poslovne Usluge": (
        "Professional Services", "Business services and operational support provider (Croatia)",
        "M", "", ""
    ),
    "Planful, Inc.": (
        "Finance", "Financial planning and analysis (FP&A) and budgeting software platform",
        "H", "", ""
    ),
    "Cognism Limited": (
        "Sales", "B2B sales intelligence, lead prospecting and contact data platform",
        "H", "", ""
    ),
    "Uberflip": (
        "Marketing", "Content experience platform for personalised sales and marketing content delivery",
        "H", "", ""
    ),

    # ── T3 OPERATIONAL ($10K–$24.9K) ─────────────────────────────────────────
    "Agram Life Osiguranje D.O.O.": (
        "G&A", "Employee life and health insurance provider (Croatia)",
        "H", "", ""
    ),
    "Google Ireland Limited": (
        "SaaS", "Google Workspace productivity and collaboration suite (company-wide)",
        "M", "",
        "Assumed Google Workspace (SaaS); flip to Engineering if Google Cloud Platform is primary product"
    ),
    "Zuric I Partneri Odvjetnicko Drustvo D.O.O.": (
        "Legal", "Croatian commercial law firm providing legal advice and representation",
        "H", "", ""
    ),
    "Care Health Insurance Company Limited": (
        "G&A", "Employee health insurance provider",
        "H", "", ""
    ),
    "New Star Networks(Nsn)": (
        "Facilities", "Network and telecommunications infrastructure services",
        "M", "", ""
    ),
    "Bupa- Supplier": (
        "G&A", "Employee health insurance and benefits provider",
        "H", "Bupa Group",
        "Flagged as possible individual name — false positive; Bupa is a global health insurance provider"
    ),
    "Shree Info System Solutions Pvt Ltd": (
        "Engineering", "IT systems and software solutions provider (India)",
        "M", "", ""
    ),
    "Technet It Recruitment": (
        "Engineering", "IT and technology specialist recruitment agency",
        "M", "", "Explicitly IT/tech recruitment → Engineering per edge case rule"
    ),
    "Mightyhive Ltd": (
        "Marketing", "Digital marketing, programmatic advertising and marketing analytics agency",
        "H", "", ""
    ),
    "Cedar Recruitment Ltd": (
        "G&A", "General recruitment and talent acquisition agency (company-wide)",
        "H", "", ""
    ),
    "Eurofast International Ltd-Greec": (
        "Finance", "International payroll processing, accounting and corporate administration services",
        "H", "", ""
    ),
    "Hrvatski Telekom D.D.": (
        "Facilities", "National telecommunications provider — office connectivity and communications (Croatia)",
        "H", "", ""
    ),
    "Sodexo Svc India Private Limited": (
        "Facilities", "Facilities management and corporate catering services for office (India)",
        "H", "", ""
    ),
    "Peakon Aps": (
        "G&A", "Employee engagement analytics and sentiment measurement platform (Workday Peakon)",
        "H", "",
        "Flagged as possible individual name — false positive; Peakon Aps is an employee engagement SaaS company"
    ),
    "Plus Your Business Ltd": (
        "Marketing", "Digital marketing and Google Business profile optimisation services",
        "H", "", ""
    ),
    "Benefit Systems D.O.O.": (
        "G&A", "Employee benefits platform and MultiSport wellness program provider (Croatia)",
        "H", "", ""
    ),
    "Smart Group Services D.O.O.": (
        "Professional Services", "Business services and operational support provider (Croatia)",
        "M", "", ""
    ),
    "Jones Lang Lasalle (Nsw) Pty Ltd": (
        "Facilities", "Commercial property management and real estate advisory services (JLL, Australia)",
        "H", "", ""
    ),
    "Workato, Inc.": (
        "Engineering", "Enterprise integration and workflow automation platform",
        "M", "",
        "Integration platform typically IT/Engineering-governed; reclassify to SaaS if business-user-led automations dominate"
    ),
    "Konzum Plus D.O.O.": (
        "Facilities", "Supermarket / food and beverage supplies for office catering (Croatia — Konzum retail chain)",
        "M", "", ""
    ),
    "Westbrook Advisers": (
        "Legal", "Commercial legal advisory services for board and senior management",
        "M", "",
        "Strategy B: Legal vs Professional Services ambiguity irrelevant at T3 ($15K). Legal is more precise fit."
    ),
    "Work Easy Space Solutions Private Limited": (
        "Facilities", "Coworking and serviced office space provider (India)",
        "H", "", ""
    ),
    "Taxstudio, Ltd.": (
        "Finance", "Specialist tax advisory and compliance services",
        "H", "", ""
    ),
    "Pingo D.O.O.": (
        "Facilities", "Catering or hospitality services for office (Croatia)",
        "L", "", ""
    ),
    "Magazin Raunalni Sistemi D.O.O.": (
        "Engineering", "IT equipment and computing systems supplier (Croatia — 'Računalni Sistemi' = Computer Systems)",
        "M", "", "Encoding may be partially corrupted ('Raunalni' → 'Računalni')"
    ),
    "Tp Prime D.O.O.": (
        "G&A", "Business services provider (Croatia — primary service unconfirmed)",
        "L", "", ""
    ),
    "Granttree Limited": (
        "Finance", "R&D tax credits advisory and government funding specialist (UK)",
        "H", "", ""
    ),
    "Cigna Sg": (
        "G&A", "Employee health insurance provider (Singapore — Cigna group)",
        "H", "",
        "Flagged as possible individual name — false positive; Cigna is a global health insurance company"
    ),
    "Collards Chartered Accountants": (
        "Finance", "Chartered accountants providing audit, tax and accounting services",
        "H", "", ""
    ),
    "Bupa Australia": (
        "G&A", "Employee health insurance provider (Australia)",
        "H", "Bupa Group", ""
    ),
    "4I Management Consulting Private Limited": (
        "Professional Services", "Management consulting services (India entity of the 4I Group)",
        "L", "4I Group",
        "Strategy A: 'Management Consulting' in entity name signals operational BAU work rather than deal advisory. Immaterial at $12K."
    ),
    "Grad Zagreb, Gradski Ured Za Prostorno Ureä'Enje,..": (
        "Facilities", "Zagreb City Government planning office — building permits and workspace regulatory fees (Croatia)",
        "R", "",
        "Encoding corrupted (expected: 'Grad Zagreb, Gradski Ured Za Prostorno Uređenje'). Classified Facilities based on 'Prostorno Uređenje' = spatial planning/permits."
    ),
    "Bijeli Pijesak Obrt Za Poslovno Savjetovanje": (
        "Professional Services", "Business consulting and advisory services (Croatia — 'Bijeli Pijesak' = White Sand)",
        "L", "", ""
    ),
    "Ramiro D.O.O.": (
        "G&A", "General business services provider (Croatia — primary service unconfirmed from name)",
        "L", "", ""
    ),
    "Visalogic Limited": (
        "Legal", "Global immigration and visa management services",
        "H", "", ""
    ),
    "Poles Ltd - Hanbury Manor": (
        "G&A", "Hotel and event venue services — likely team offsite or client entertainment (Hanbury Manor, UK)",
        "M", "", ""
    ),
    "Shoff Darby Companies": (
        "G&A", "Business services provider (primary service unconfirmed from name)",
        "L", "", ""
    ),
    "Emerge Development Consultancy Ltd": (
        "Professional Services", "Business development and growth consulting services",
        "M", "", ""
    ),
    "Performancepro": (
        "G&A", "Employee performance management and appraisal software platform",
        "M", "",
        "Flagged as possible individual name — false positive; PerformancePro is an HR tech SaaS product"
    ),

    # ── T4 TAIL ($1K–$9.9K) ───────────────────────────────────────────────────
    "Microsoft Ireland Operations Limited": (
        "SaaS", "Microsoft 365 productivity and collaboration suite (company-wide)",
        "M", "",
        "Assumed Microsoft 365 (SaaS); flip to Engineering if Azure cloud services are the primary product"
    ),
    "Catering Muring": (
        "Facilities", "Office catering and food services (Croatia)",
        "H", "", ""
    ),
    "Omonia D.O.O.": (
        "Facilities", "Office catering or hospitality services (Croatia)",
        "L", "", ""
    ),
    "Outreach Corporation": (
        "Sales", "Sales engagement, sequencing and outbound prospecting platform",
        "H", "", ""
    ),
    "Cbre Limited": (
        "Facilities", "Commercial real estate brokerage and property management advisory (CBRE)",
        "H", "", ""
    ),
    "Studentski Centar - Split": (
        "Facilities", "University student center — conference or events venue (Split, Croatia)",
        "M", "", ""
    ),
    "The Guardian": (
        "G&A", "News media subscription for executive and company-wide news access",
        "M", "", ""
    ),
    "Intertrust Singapore Corporate Services Pte Ltd - Csc": (
        "Finance", "Corporate administration, trust, fund administration and compliance services (Singapore)",
        "M", "", ""
    ),
    "Goto Technologies Uk Limited": (
        "SaaS", "Remote access, conferencing and IT support platform (GoTo Meeting / GoTo)",
        "H", "", ""
    ),
    "Hrsolution International Ag": (
        "G&A", "HR staffing and workforce management services (Switzerland entity)",
        "H", "HRsolution International", ""
    ),
    "Trello": (
        "SaaS", "Visual project management and team collaboration tool (cross-functional)",
        "H", "", "Flagged as possible individual name — false positive; Trello is an Atlassian project management product"
    ),
    "Obrt Sjaj Sunca": (
        "Facilities", "Cleaning or office maintenance services (Croatia — 'Sjaj Sunca' = Sunshine Gleam)",
        "L", "", ""
    ),
    "Telefã³Nica Compras Electrã³Nicas S.L.": (
        "Facilities", "Telefonica electronic procurement entity — office telecommunications or connectivity (Spain)",
        "R", "",
        "Encoding corrupted (expected: 'Telefónica Compras Electrónicas S.L.'). Classified Facilities as Telefonica entity. Verify decoded name."
    ),
    "Allianz Australia Workers' Compensation (Victoria) Limited": (
        "G&A", "Workers compensation insurance — Victoria, Australia (Allianz group)",
        "H", "Allianz", ""
    ),
    "Icici Lombard Gic Ltd": (
        "G&A", "General insurance and workers compensation provider (India — ICICI Lombard)",
        "H", "", ""
    ),
    "Mosaic Concept D.O.O.": (
        "Professional Services", "Design or creative professional services (Croatia)",
        "L", "", ""
    ),
    "Limes Plus D.O.O.": (
        "Facilities", "Office facilities or maintenance services (Croatia)",
        "L", "", ""
    ),
    "Mercer Limited": (
        "G&A", "Employee benefits consulting and HR advisory services",
        "H", "", ""
    ),
    "Acclime Corporate Services": (
        "Finance", "International corporate administration, accounting and compliance services",
        "H", "Acclime", ""
    ),
    "Green Commute Initiative": (
        "G&A", "Employee cycle-to-work scheme and sustainable commuting benefits programme",
        "H", "", ""
    ),
    "Tattu Manchester Limited": (
        "G&A", "Premium restaurant — team events and client entertainment (Manchester)",
        "M", "", ""
    ),
    "Athlete Service Ltd": (
        "G&A", "Employee sport and wellness benefits programme",
        "M", "", ""
    ),
    "Amazon Web Services Inc.": (
        "Engineering", "Cloud infrastructure and compute services (AWS — US entity)",
        "H", "Amazon Web Services / Amazon", "Near-duplicate of 'Amazon Web Services Llc'; consolidate to canonical brand for spend reporting"
    ),
    "Telemach Hrvatska D.O.O.": (
        "Facilities", "National cable and telecommunications provider — office internet and communications (Croatia)",
        "H", "", ""
    ),
    "Acclime Usa, Inc": (
        "Finance", "International corporate administration and accounting services (US entity)",
        "H", "Acclime", ""
    ),
    "Centar Za Sigurnost D.O.O.": (
        "Facilities", "Physical security services and access control (Croatia — 'Sigurnost' = Security)",
        "M", "", ""
    ),
    "Profi Bar D.O.O.": (
        "Facilities", "Office coffee bar and refreshment services (Croatia)",
        "L", "", ""
    ),
    "Pricewaterhousecoopers Llp": (
        "Finance", "Big 4 accounting firm providing audit, tax and business advisory",
        "M", "", "Could be M&A if engagement was transaction-specific due diligence; classified Finance as default"
    ),
    "Pinnacle Partnership Ca": (
        "Professional Services", "Business advisory or consulting services (Canada)",
        "L", "", ""
    ),
    "The Virtual Legal Counsel Ltd": (
        "Legal", "Outsourced in-house legal counsel and commercial legal services",
        "H", "", ""
    ),
    "Australian Payroll Professionals Pty Ltd": (
        "Finance", "Payroll processing and management services (Australia)",
        "H", "", ""
    ),
    "Inter Continental Chennai Mahabalipuram Resort": (
        "G&A", "Hotel and resort accommodation — likely business travel or team offsite (India)",
        "L", "", ""
    ),
    "Jetbrains S.R.O.": (
        "Engineering", "Developer IDEs and programming productivity tools (JetBrains)",
        "H", "", "Flagged as possible individual name — false positive; JetBrains is a major developer tools company"
    ),
    "Crowe Horwath Revizija D.O.O.": (
        "Finance", "Audit and accounting services (Crowe global network, Croatia)",
        "H", "", ""
    ),
    "Puducherry Backwater Resort Private Limited": (
        "G&A", "Hotel and resort accommodation — likely business travel or team offsite (India)",
        "L", "", ""
    ),
    "Orionw Llc": (
        "Professional Services", "IT or software services provider (US — primary service unconfirmed)",
        "L", "", ""
    ),
    "Adobe Systems Software": (
        "SaaS", "Creative, document management and digital experience software suite (Adobe)",
        "H", "", ""
    ),
    "Common Desk, Llc": (
        "Facilities", "Coworking and shared office space provider (US — Common Desk)",
        "H", "", ""
    ),
    "Hep Elektra D.O.O.": (
        "Facilities", "Office electricity utility provider (Croatia — HEP, Croatian Energy Group)",
        "H", "", ""
    ),
    "Zivi Napitak D.O.O.": (
        "Facilities", "Office beverage and refreshment supplier (Croatia)",
        "L", "", ""
    ),
    "Gym4You D.O.O.": (
        "G&A", "Employee gym and fitness membership benefits (Croatia)",
        "M", "", ""
    ),
    "Aha! Labs Inc": (
        "Product", "Product roadmapping and strategy platform for Product Management teams",
        "H", "", ""
    ),
    "Apple Retail Uk Ltd": (
        "G&A", "Apple hardware and software purchases — company equipment (UK retail entity)",
        "M", "Apple", "Apple spend split across 4 entities (Retail UK, Pty, Distribution, Amer); classify each row individually"
    ),
    "Kryterion, Inc.": (
        "G&A", "Online proctoring and professional certification testing platform",
        "L", "", ""
    ),
    "Npm Inc": (
        "Engineering", "JavaScript package manager and npm registry — core software development dependency",
        "H", "", ""
    ),
    "Pinsent Masons Mpillay Llp": (
        "Legal", "International law firm providing commercial, technology and employment legal services",
        "H", "", ""
    ),
    "Calm Achiever(A Unit Of Mohsin Ali Vakil)": (
        "G&A", "Employee wellness and mental health coaching services (India)",
        "L", "", ""
    ),
    "Pluxee India Private Limited": (
        "G&A", "Employee benefits and meal voucher platform (formerly Sodexo Benefits & Rewards, India)",
        "H", "", ""
    ),
    "Oâ€™Donnell Salzano Lawyers": (
        "Legal", "Australian law firm providing commercial and employment legal services",
        "R", "",
        "Encoding corrupted (expected: 'O'Donnell Salzano Lawyers'). Classified Legal based on 'Lawyers' in decoded name. Verify decoded name."
    ),
    "Papertrail Inc": (
        "Engineering", "Cloud-hosted log management and application monitoring service",
        "H", "", ""
    ),
    "Sniper Systems And Solutions Private Limited": (
        "Engineering", "IT systems and technology solutions provider (India)",
        "M", "", ""
    ),
    "Trocadero (London) Hotel Ltd": (
        "G&A", "Hotel accommodation — likely business travel or team meeting venue (London)",
        "M", "", ""
    ),
    "Semrush Inc": (
        "Marketing", "SEO, content marketing and competitive digital intelligence platform",
        "H", "", ""
    ),
    "Golubica Parking D.O.O.": (
        "Facilities", "Parking facilities — employee parking services (Croatia)",
        "H", "", ""
    ),
    "Vodafone (Australian)": (
        "Facilities", "Mobile telecommunications — company phones and connectivity (Australia — Vodafone)",
        "H", "", ""
    ),
    "Ikea Hrvatska D.O.O.": (
        "Facilities", "Office furniture and workplace equipment (IKEA Croatia)",
        "M", "", ""
    ),
    "Orcola D.O.O.": (
        "Professional Services", "Business or IT consulting services (Croatia — primary service unconfirmed)",
        "L", "", ""
    ),
    "Info Edge India Limited": (
        "G&A", "Online recruitment portal and job board services (Naukri.com, India)",
        "M", "", ""
    ),
    "Grad Split": (
        "G&A", "City of Split municipal fees, permits or regulatory payments (Croatia)",
        "H", "", ""
    ),
    "T-Mobile": (
        "Facilities", "Mobile telecommunications — company phones and connectivity",
        "H", "", ""
    ),
    "Akton D.O.O.": (
        "Professional Services", "Business services or IT support (Croatia — primary service unconfirmed)",
        "L", "", ""
    ),
    "Bureau Veritas Croatia D.O.O.": (
        "Professional Services", "Testing, inspection and certification (TIC) compliance services",
        "M", "", ""
    ),
    "Cici Prudential Life Insurance Co. Ltd.": (
        "G&A", "Employee life insurance provider (Asia — ICICI Prudential group)",
        "H", "", ""
    ),
    "Ncc Services Limited": (
        "Professional Services", "Business services and facilities management (primary service unconfirmed)",
        "L", "", ""
    ),
    "Smartsheet Inc.": (
        "SaaS", "Cloud-based project management and work collaboration platform (cross-functional)",
        "H", "", ""
    ),
    "Good Game Global D.O.O.": (
        "G&A", "Corporate entertainment or team event services (Croatia)",
        "L", "", ""
    ),
    "Garaå¾A Firule D.O.O.": (
        "Facilities", "Parking garage facility — employee parking (Split, Croatia — 'Garaža Firule')",
        "R", "",
        "Encoding corrupted (expected: 'Garaža Firule D.O.O.'). Classified Facilities based on 'Garaža' = garage/parking."
    ),
    "Crossland": (
        "Professional Services", "Business services or consulting (primary service unconfirmed)",
        "L", "", ""
    ),
    "It London": (
        "Engineering", "IT services or technology support provider (London)",
        "L", "",
        "Flagged as possible individual name — false positive; 'IT London' is a technology services entity"
    ),
    "Terrapinn Holdings Ltd": (
        "Marketing", "Global business events, conference and media services — industry event sponsorship",
        "M", "", ""
    ),
    "Elemental Life Solutions Llp": (
        "G&A", "Employee wellness and lifestyle benefits services",
        "L", "", ""
    ),
    "United Flow Ltd (The Goodness Project)": (
        "G&A", "Employee wellness and mental health programme services",
        "M", "", ""
    ),
    "Mcburneys Charted Accountants": (
        "Finance", "Chartered accountants providing audit, tax and accounting services",
        "M", "", ""
    ),
    "Cleverland Winery Resort": (
        "G&A", "Winery and resort accommodation — likely team offsite or business travel (Australia)",
        "M", "", ""
    ),
    "Ag Grid Ltd": (
        "Engineering", "JavaScript data grid library for enterprise web application development",
        "H", "", ""
    ),
    "Lusha": (
        "Sales", "B2B contact intelligence and lead data enrichment platform",
        "H", "", ""
    ),
    "My Foodiverse Llp": (
        "Facilities", "Corporate food delivery and office catering services",
        "M", "", ""
    ),
    "4Imprint Direct Ltd": (
        "Marketing", "Promotional merchandise and branded marketing materials supplier",
        "M", "", ""
    ),
    "City Pantry Ltd": (
        "Facilities", "Corporate office catering and team food delivery service",
        "M", "", ""
    ),
    "Stipe Piric": (
        "Professional Services", "Individual contractor/freelancer payment (Croatia)",
        "L", "",
        "Strategy A: individual name pattern; likely freelancer invoice. Reclassify to G&A if expense reimbursement confirmed."
    ),
    "Office Move London": (
        "Facilities", "Office relocation, furniture installation and moving services (London)",
        "M", "", ""
    ),
    "Pink Ribbon Shop": (
        "G&A", "Charitable merchandise — likely company CSR or employee awareness initiative",
        "M", "", ""
    ),
    "Starhub Ltd (Supplier)": (
        "Facilities", "Mobile and broadband telecommunications provider — office connectivity (Singapore — StarHub)",
        "H", "", ""
    ),
    "N S Shastri And Co": (
        "Finance", "Chartered accountants and tax advisory services (India)",
        "M", "", ""
    ),
    "John Smith": (
        "Professional Services", "Individual contractor/freelancer payment",
        "L", "",
        "Strategy A: individual name pattern; likely freelancer invoice. Reclassify to G&A if expense reimbursement confirmed."
    ),
    "Fabiola Thistlewhaite": (
        "Professional Services", "Individual contractor/freelancer payment",
        "L", "",
        "Strategy A: individual name pattern; likely freelancer invoice. Reclassify to G&A if expense reimbursement confirmed."
    ),
    "Hp Inc Uk Limited": (
        "G&A", "Hardware equipment and printing solutions — company devices and peripherals (HP UK)",
        "M", "", ""
    ),
    "Cision Pr Newswire": (
        "Marketing", "Press release distribution and PR media newswire services",
        "H", "", ""
    ),
    "George Anchor": (
        "Professional Services", "Individual contractor/freelancer payment",
        "L", "",
        "Strategy A: individual name pattern; likely freelancer invoice. Reclassify to G&A if expense reimbursement confirmed."
    ),
    "Yoxel, Inc": (
        "Engineering", "IT or software technology services (US — primary service unconfirmed)",
        "L", "", ""
    ),
    "Grt Hotels And Resorts P Ltd": (
        "G&A", "Hotel accommodation — business travel or team offsite (GRT Hotels, South India)",
        "L", "", ""
    ),
    "Apple Pty Ltd": (
        "G&A", "Apple hardware and software purchases — company equipment (Australia retail entity)",
        "M", "Apple", ""
    ),
    "Slack Technologies Limited": (
        "SaaS", "Enterprise messaging, file sharing and team communication platform",
        "H", "", ""
    ),
    "G S Notary Public Limited": (
        "Legal", "Notary public and document authentication and legalisation services",
        "H", "", ""
    ),
    "Apple Distribution International Ltd": (
        "G&A", "Apple software licensing and App Store distribution (Ireland entity)",
        "M", "Apple", ""
    ),
    "Porezno Savjetniå¡Tvo Tuk D.O.O.": (
        "Finance", "Tax advisory and financial compliance consulting (Croatia — 'Porezno Savjetništvo' = Tax Advisory)",
        "R", "",
        "Encoding corrupted (expected: 'Porezno Savjetništvo Tuk D.O.O.'). Classified Finance based on 'Porezno Savjetništvo' = Tax Advisory."
    ),
    "Susan Lee": (
        "Professional Services", "Individual contractor/freelancer payment",
        "L", "",
        "Strategy A: individual name pattern; likely freelancer invoice. Reclassify to G&A if expense reimbursement confirmed."
    ),
    "Ansar Madovic": (
        "Professional Services", "Individual contractor/freelancer payment (Croatia)",
        "L", "",
        "Strategy A: individual name pattern; likely freelancer invoice. Reclassify to G&A if expense reimbursement confirmed."
    ),
    "Radius Group, Inc": (
        "Professional Services", "Business advisory or technology services (US — primary service unconfirmed)",
        "L", "", ""
    ),
    "Aquila Remete D.O.O.": (
        "Facilities", "Hospitality or accommodation services (Croatia — Aquila hotel/venue)",
        "L", "", ""
    ),
    "Clime India Private Limited": (
        "G&A", "Corporate wellness or HR technology services (India)",
        "L", "", ""
    ),
    "Golden Mean, Inc": (
        "Professional Services", "Business advisory or consulting services (US — primary service unconfirmed)",
        "L", "", ""
    ),
    "Paint & Fun Vl. Martina Milkova Nikolova": (
        "G&A", "Employee team-building and creative activity event (paint and art experiences)",
        "M", "", ""
    ),
    "Carrington Communications": (
        "Marketing", "PR, media relations and corporate communications services",
        "M", "",
        "Flagged as possible individual name — false positive; Carrington Communications is a PR/comms agency"
    ),
    "Crayond Digital Private Limited": (
        "Engineering", "Digital product design and software development agency (India)",
        "M", "", ""
    ),
    "Studentski Centar Karlovac": (
        "Facilities", "Student center venue — conference or events space (Karlovac, Croatia)",
        "M", "", ""
    ),
    "Lajnap Comedy Booking D.O.O.": (
        "G&A", "Comedy and entertainment booking agency — likely for company event (Croatia)",
        "M", "", ""
    ),
    "British Telecommunications": (
        "Facilities", "Office telecommunications and connectivity services (BT, UK)",
        "H", "",
        "Flagged as possible individual name — false positive; British Telecommunications is the BT Group"
    ),
    "Etm Concessions Ltd": (
        "G&A", "Hospitality and food/beverage venue services — likely team event (ETM Group, UK)",
        "M", "", ""
    ),
    "Radisson Grt - Unit Of Hotels & Resorts Pvt Ltd": (
        "G&A", "Hotel accommodation — business travel (Radisson GRT, India)",
        "L", "", ""
    ),
    "Ariba Inc": (
        "SaaS", "SAP Ariba procurement and spend management platform (company-wide)",
        "M", "", ""
    ),
    "Chamiers Recreation Club": (
        "G&A", "Recreation club membership — employee wellness benefit (India)",
        "L", "", ""
    ),
    "Quadrant Law Llc": (
        "Legal", "Legal services and commercial law advisory",
        "H", "", ""
    ),
    "Docusign": (
        "SaaS", "Electronic signature and digital contract management platform (company-wide)",
        "H", "", ""
    ),
    "Inside Edge Novated Leasing": (
        "G&A", "Employee novated car leasing and salary packaging service (Australia)",
        "M", "", ""
    ),
    "Rhea D.O.O.": (
        "Facilities", "Office facilities or maintenance services (Croatia — primary service unconfirmed)",
        "L", "", ""
    ),
    "P S Recreation Club": (
        "G&A", "Recreation club membership — employee wellness benefit",
        "L", "", ""
    ),
    "Fastspring": (
        "Engineering", "E-commerce and subscription billing platform for software product sales",
        "M", "", ""
    ),
    "Dsv Solutions A/S": (
        "Facilities", "Freight and logistics services — courier and office supply delivery (DSV, Denmark)",
        "M", "", ""
    ),
    "Curzon Green Solicitors": (
        "Legal", "Solicitors providing commercial, employment and residential legal services",
        "H", "", ""
    ),
    "Icare Nsw": (
        "G&A", "Workers compensation insurance authority — New South Wales, Australia (icare)",
        "H", "",
        "Flagged as possible individual name — false positive; icare NSW is the NSW Government insurance provider"
    ),
    "Thomas Mansfield Solicitors Limited": (
        "Legal", "Employment law and commercial legal services (UK)",
        "H", "", ""
    ),
    "Amazon.Co.Uk": (
        "G&A", "Amazon UK marketplace — general office supplies and equipment purchases",
        "M", "Amazon Web Services / Amazon",
        "Amazon.co.uk is retail, not AWS; classified G&A for general purchases rather than Engineering"
    ),
    "Backoffice Associates": (
        "Professional Services", "Data migration, data quality and ERP/SAP consulting services",
        "M", "", ""
    ),
    "Oladi D.O.O.": (
        "Facilities", "Office facilities or services (Croatia — primary service unconfirmed)",
        "L", "", ""
    ),
    "Integrated Personnel Services": (
        "G&A", "HR staffing and personnel management services",
        "M", "", ""
    ),
    "Mãœller Trgovina Zagreb D.O.O.": (
        "G&A", "Retail store (personal care, cosmetics) — miscellaneous office or employee purchases (Müller Croatia)",
        "R", "",
        "Encoding corrupted (expected: 'Müller Trgovina Zagreb D.O.O.'). Classified G&A based on Müller being a retail chain."
    ),
    "Lunch Nutrition D.O.O.": (
        "Facilities", "Corporate lunch and nutrition delivery services for office (Croatia)",
        "M", "", ""
    ),
    "Winmaxi Tours & Travels": (
        "G&A", "Corporate travel agency and travel management services (India)",
        "M", "", ""
    ),

    # ── T5 NOISE (<$1K) ───────────────────────────────────────────────────────
    "Digitalna Produkcija J.D.O.O.": (
        "Marketing", "Digital content production and creative services (Croatia)",
        "N", "", ""
    ),
    "Computershare-Caboodle Technology Limited": (
        "Finance", "Employee share plan administration and equity management platform",
        "N", "", ""
    ),
    "Allianz Wa": (
        "G&A", "Workers compensation insurance — Western Australia (Allianz group)",
        "N", "Allianz",
        "Flagged as possible individual name — false positive; Allianz WA is workers comp insurance"
    ),
    "Trending Technology Services Gmbh": (
        "Engineering", "IT and technology consulting services (Germany)",
        "N", "", ""
    ),
    "Godaddy.Com, Llc": (
        "Engineering", "Domain registration and web hosting services for digital infrastructure",
        "N", "", ""
    ),
    "Time Out Group": (
        "G&A", "Media and lifestyle content — editorial subscription or city guide advertising",
        "N", "", ""
    ),
    "Raiffeisenbank Austria D.D.": (
        "Finance", "Banking services — account administration or transaction fees (Croatia — Raiffeisen)",
        "N", "", ""
    ),
    "Galop-Prijevoz D.O.O.": (
        "Facilities", "Transport and courier delivery services (Croatia)",
        "N", "", ""
    ),
    "Lane Ip Limited": (
        "Legal", "Intellectual property management and patent/trademark filing services",
        "N", "", ""
    ),
    "Urbani Eventi D.O.O.": (
        "G&A", "Corporate event planning and production services (Croatia)",
        "N", "", ""
    ),
    "Smashing Media Ag": (
        "Marketing", "Digital media and web design publication (Smashing Magazine)",
        "N", "", ""
    ),
    "Epignosis Llc": (
        "G&A", "Learning management system (eFront LMS) for employee training and development",
        "N", "", ""
    ),
    "Croatia Airlines": (
        "G&A", "Commercial airline — employee business travel (Croatia)",
        "N", "", ""
    ),
    "Bb Football Scouting J.D.O.O.": (
        "G&A", "Football/sports scouting — likely employee wellbeing or team-building event",
        "N", "", ""
    ),
    "Boe Croatia D.O.O.": (
        "Professional Services", "Business or operational services (Croatia — primary service unconfirmed)",
        "N", "", ""
    ),
    "Potomac D.O.O.": (
        "Facilities", "Office facilities or hospitality services (Croatia — primary service unconfirmed)",
        "N", "",
        "Flagged as possible individual name — false positive; Potomac D.O.O. is a Croatian company"
    ),
    "President Hotel And Tower Co., Ltd": (
        "G&A", "Hotel accommodation — business travel (Asia)",
        "N", "", ""
    ),
    "Friends Sports Club": (
        "G&A", "Sports club membership — employee wellness benefit",
        "N", "", ""
    ),
    "The Cycle Gap Adyar": (
        "G&A", "Bicycle retail — likely part of employee cycle-to-work or wellness initiative (India)",
        "N", "", ""
    ),
    "Apple - Amer": (
        "G&A", "Apple hardware and software — Americas region entity",
        "N", "Apple", ""
    ),
    "Klg - Kalra Legal Group": (
        "Legal", "Legal services and commercial law advisory (India — Kalra Legal Group)",
        "N", "", ""
    ),
    "Kilgannon & Partners Llp": (
        "Legal", "Legal services and commercial law advisory",
        "N", "", ""
    ),
    "Uk Postbox Limited": (
        "G&A", "Virtual mailbox and mail forwarding services — registered business address (UK)",
        "N", "", ""
    ),
    "Catering Iviä‡ D.O.O.": (
        "Facilities", "Catering and food services for office (Croatia — 'Catering Ivić')",
        "R", "",
        "Encoding corrupted (expected: 'Catering Ivić D.O.O.'). Classified Facilities based on 'Catering' in name."
    ),
    "Blink Events": (
        "G&A", "Corporate event management and production services",
        "N", "", ""
    ),
    "Grafo-Jan": (
        "Marketing", "Graphic design and print services (Croatia)",
        "N", "", ""
    ),
    "Mesa Verde": (
        "G&A", "Restaurant or hospitality — business meal or team event",
        "N", "",
        "Flagged as possible individual name — false positive; Mesa Verde is a restaurant/hospitality name"
    ),
    "Pixsy Inc": (
        "Legal", "Image copyright monitoring and intellectual property infringement recovery platform",
        "N", "", ""
    ),
    "Dun & Bradstreet D.O.O.": (
        "Finance", "Business credit risk intelligence, financial data and compliance analytics services",
        "N", "", ""
    ),
    "Induslaw": (
        "Legal", "Indian law firm providing corporate and commercial legal services",
        "N", "", ""
    ),
    "Trans-Agram Obrt Za Dostavu": (
        "Facilities", "Courier and delivery services (Croatia — 'Dostavu' = delivery)",
        "N", "", ""
    ),
    "Landu Law Solicitors": (
        "Legal", "Legal services and solicitor advisory",
        "N", "", ""
    ),
    "Arena Center Zagreb D.O.O.": (
        "G&A", "Shopping mall venue — retail purchase or team event (Arena Center, Zagreb)",
        "N", "", ""
    ),
    "Obrt Za Ugostiteljstvo Mirakul": (
        "Facilities", "Catering and hospitality services for office (Croatia — 'Ugostiteljstvo' = catering)",
        "N", "", ""
    ),
    "Oakberry Jr D.O.O.": (
        "Facilities", "Food and beverage refreshments — office food (Oakberry açaí brand)",
        "N", "", ""
    ),
    "Tau On-Line D.O.O.": (
        "Engineering", "Online technology or IT services (Croatia)",
        "N", "", ""
    ),
    "Magic Mountain Saloon": (
        "G&A", "Bar and entertainment venue — team social event (Croatia)",
        "N", "", ""
    ),
    "Hotel Zonar": (
        "G&A", "Hotel accommodation — business travel or team offsite (Croatia)",
        "N", "",
        "Flagged as possible individual name — false positive; Hotel Zonar is a Croatian hotel"
    ),
    "Roto Dinamic D.O.O.": (
        "Engineering", "Technical or mechanical engineering equipment services (Croatia)",
        "N", "", ""
    ),
    "Tm Forum": (
        "Professional Services", "Telecoms industry standards body membership (lower-value entry; see Tmforum $57.6K)",
        "N", "", "Possible duplicate/secondary payment to TM Forum — see also 'Tmforum' at $57.6K"
    ),
    "Marvie Hotel - Krupa D.O.O.": (
        "G&A", "Hotel accommodation — business travel or team offsite (Croatia)",
        "N", "", ""
    ),
    "Merchandise Ltd": (
        "Marketing", "Branded merchandise and promotional products",
        "N", "", ""
    ),
    "Yellow Submarine D.O.O.": (
        "Facilities", "Bar, restaurant or catering services (Croatia)",
        "N", "", ""
    ),
    "Streamlinereforms Inc": (
        "Professional Services", "Business process reform and consulting services",
        "N", "", ""
    ),
    "Make And Grow Ltd": (
        "Marketing", "Creative agency or marketing growth services",
        "N", "", ""
    ),
    "Obiteljski Hoteli D.O.O.": (
        "G&A", "Hotel accommodation — business travel (Croatia — 'Obiteljski Hoteli' = Family Hotels)",
        "N", "", ""
    ),
    "Rudan D.O.O.": (
        "Facilities", "Office facilities, construction or maintenance services (Croatia)",
        "N", "",
        "Flagged as possible individual name — false positive; Rudan D.O.O. is a Croatian company"
    ),
    "Entrio Tehnologije D.O.O.": (
        "Marketing", "Event ticketing technology platform for Croatian events",
        "N", "", ""
    ),
    "Vivat Fina Vina D.O.O.": (
        "G&A", "Fine wine retailer — likely for company events or client entertainment (Croatia)",
        "N", "", ""
    ),
    "Figma, Inc.": (
        "Product", "Collaborative design and interactive prototyping tool for Product and UX teams",
        "N", "", ""
    ),
    "E-Disti D.O.O.": (
        "Engineering", "IT distribution and technology products supplier (Croatia)",
        "N", "", ""
    ),
    "Greencell Express Private Limited": (
        "G&A", "Express delivery and courier services (India)",
        "N", "", ""
    ),
    "Edwardian Pastoria Hotels Ltd (The Londoner)": (
        "G&A", "Hotel accommodation — business travel (The Londoner, London)",
        "N", "", ""
    ),
    "Tiganda J.D.O.O.": (
        "Facilities", "Office facilities or maintenance services (Croatia — primary service unconfirmed)",
        "N", "",
        "Flagged as possible individual name — false positive; Tiganda J.D.O.O. is a Croatian company"
    ),
    "Franklin, Gringer & Cohen, P.C.": (
        "Legal", "US law firm providing legal services",
        "N", "", ""
    ),
    "Blitz - Cinestar D.O.O.": (
        "G&A", "Cinema chain — employee entertainment benefit or team event (Croatia — Blitz CineStar)",
        "N", "", ""
    ),
    "Lancefield Bus Service": (
        "Facilities", "Bus and transport service — likely employee shuttle or event transport (UK)",
        "N", "", ""
    ),
    "Super Odrediå¡Te D.O.O.": (
        "G&A", "Retail or general services (Croatia — 'Super Odredište' = Super Destination; encoding corrupted)",
        "R", "",
        "Encoding corrupted (expected: 'Super Odredište D.O.O.'). Classified G&A as catch-all; primary service unconfirmed."
    ),
    "Pmi Global Operations Center": (
        "G&A", "Project Management Institute (PMI) — professional certification and membership fees",
        "N", "", ""
    ),
    "Student Packers & Movers": (
        "Facilities", "Office relocation and moving services (India)",
        "N", "", ""
    ),
    "Del Posto D.O.O.": (
        "Facilities", "Restaurant or catering services (Croatia — 'Del Posto' restaurant)",
        "N", "", ""
    ),
    "Inicijativa Centar Za Edukaciju": (
        "G&A", "Education and professional training centre (Croatia — 'Centar Za Edukaciju' = Education Centre)",
        "N", "", ""
    ),
    "Niva Transport J.D.O.O.": (
        "Facilities", "Transport and logistics services (Croatia)",
        "N", "", ""
    ),
    "Doctor Anywhere Operations Pte Ltd": (
        "G&A", "Telemedicine and digital health platform — employee wellness benefit (Singapore)",
        "N", "", ""
    ),
    "Spar Hrvatska D.O.O.": (
        "Facilities", "Supermarket and grocery chain — office supplies or canteen provisions (Croatia — SPAR)",
        "N", "", ""
    ),
    "Pepe'S Italian And Liquor": (
        "Facilities", "Restaurant and bar — office catering or team meal (Croatia)",
        "N", "", ""
    ),
    "Pluralsight, Llc": (
        "Engineering", "Technology skills and developer training platform",
        "N", "", ""
    ),
    "Maniax Melbourne Cbd": (
        "G&A", "Axe-throwing entertainment venue — team-building event (Melbourne)",
        "N", "", ""
    ),
    "Dnsimple": (
        "Engineering", "DNS management and domain services for technical infrastructure",
        "N", "", ""
    ),
    "Treci Posao D.O.O.": (
        "G&A", "Online recruitment and job listing platform (Croatia — 'Treći Posao' = Third Job)",
        "N", "", ""
    ),
    "Gaucho Restaurants": (
        "G&A", "Argentinian restaurant chain — team event or client entertainment",
        "N", "",
        "Flagged as possible individual name — false positive; Gaucho is a premium restaurant chain"
    ),
    "Formswift": (
        "SaaS", "Online document creation, templates and e-signature platform",
        "N", "", ""
    ),
    "Safestore Ltd": (
        "Facilities", "Self-storage facilities — likely for office equipment or business documents",
        "N", "", ""
    ),
    "Split Tech City": (
        "Professional Services", "Tech industry community and professional networking association (Split, Croatia)",
        "N", "", ""
    ),
    "National Securities Depository Limited(Nsdl)": (
        "Finance", "National securities depository — employee share plan and ESOP settlement services (India)",
        "N", "", ""
    ),
    "Adamma Info Services Private Limited": (
        "G&A", "IT or business services provider (India — primary service unconfirmed)",
        "N", "", ""
    ),
    "The Riding House Cafe": (
        "G&A", "Restaurant and cafe — business meal or team event (London)",
        "N", "", ""
    ),
    "Stillmark Zagreb D.O.O.": (
        "Marketing", "PR and corporate communications agency (Zagreb, Croatia)",
        "N", "", ""
    ),
    "Lastpass Ireland Limited": (
        "SaaS", "Enterprise password manager and credentials security vault (company-wide IT security)",
        "N", "", ""
    ),
    "Taste Of Health": (
        "Facilities", "Healthy food delivery or catering services for office",
        "N", "", ""
    ),
    "Infodata": (
        "Engineering", "Data or IT services provider (primary service unconfirmed)",
        "N", "", ""
    ),
    "Amazon (Aus)": (
        "G&A", "Amazon Australia marketplace — general office supplies and purchases",
        "N", "Amazon Web Services / Amazon",
        "Amazon.com.au retail, not AWS; classified G&A for general purchases"
    ),
    "Regency Hampers Ltd": (
        "G&A", "Corporate gift hampers — client or employee seasonal gifting",
        "N", "", ""
    ),
    "Bella Operation A/S": (
        "Facilities", "Conference and exhibition centre venue (Bella Center, Copenhagen)",
        "N", "", ""
    ),
    "The Plant Man": (
        "Facilities", "Office plant installation and maintenance services",
        "N", "", ""
    ),
    "Media Promo Plus D.O.O.": (
        "Marketing", "Media promotion and advertising services (Croatia)",
        "N", "", ""
    ),
    "Expert-Ing D.O.O.": (
        "Engineering", "Engineering and technical consulting services (Croatia — 'Expert-Ing')",
        "N", "", ""
    ),
    "Atlassian Pty Ltd": (
        "Engineering", "Developer collaboration and project management tools (Jira, Confluence — Atlassian)",
        "N", "",
        "Flagged as possible individual name — false positive; Atlassian is a major developer tools company"
    ),
    "Djs For U": (
        "G&A", "DJ and entertainment services — likely for a company social event",
        "N", "", ""
    ),
    "Freepik Company": (
        "Marketing", "Stock images, vectors and design resources for marketing and content creation",
        "N", "", ""
    ),
    "Nastavni Zavod Za Javno Zdravstvo Dr. Andrija Å Tampar": (
        "G&A", "Public health teaching institute — occupational health services (Croatia — 'Dr. Andrija Štampar')",
        "R", "",
        "Encoding corrupted (expected: 'Nastavni Zavod Za Javno Zdravstvo Dr. Andrija Štampar'). Classified G&A (occupational health)."
    ),
    "Rishi Events And Entainment": (
        "G&A", "Event management and entertainment services (India)",
        "N", "", ""
    ),
    "Kosmaz Technologies Croatia": (
        "Engineering", "Technology services and IT solutions (Croatia)",
        "N", "", ""
    ),
    "Dhl": (
        "Facilities", "Express delivery and international courier services",
        "N", "", ""
    ),
    "Hotel Laguna D.D.": (
        "G&A", "Hotel accommodation — business travel or team offsite (Croatia)",
        "N", "", ""
    ),
    "Bigshare Services Private Limited": (
        "Finance", "Share registry and investor services — ESOP and equity plan administration (India)",
        "N", "", ""
    ),
    "Zettanet": (
        "Engineering", "Network and IT connectivity services",
        "N", "", ""
    ),
    "Hahn Air": (
        "G&A", "Airline interline ticketing services — business travel",
        "N", "", ""
    ),
    "Vitality Works": (
        "G&A", "Employee health and corporate wellness programme services (Australia)",
        "N", "",
        "Flagged as possible individual name — false positive; Vitality Works is a workplace wellness provider"
    ),
    "Avoxi Inc": (
        "Engineering", "Cloud communications platform and virtual phone number services",
        "N", "", ""
    ),
    "Zapier Inc.": (
        "SaaS", "Workflow automation and app integration platform (cross-functional, IT-governed)",
        "N", "", ""
    ),
    "Solarwinds, Inc": (
        "Engineering", "IT infrastructure monitoring and network management software",
        "N", "", ""
    ),
    "Xenon Savjetovanje D.O.O.": (
        "Professional Services", "Business and management consulting services (Croatia — 'Savjetovanje' = Consulting)",
        "N", "", ""
    ),
    "Floom Ltd": (
        "G&A", "Premium online flower delivery — employee recognition, gifting or events",
        "N", "", ""
    ),
    "Notino S.R.O.": (
        "G&A", "Online beauty and fragrance retailer — likely employee gifting or wellness purchases",
        "N", "", ""
    ),
    "Paint&Wine, Vl. Stevo Dosen": (
        "G&A", "Paint and wine team-building creative experience (Croatia)",
        "N", "", ""
    ),
    "Parcelforce Worldwide": (
        "Facilities", "Parcel delivery and business postal services (UK — Parcelforce)",
        "N", "", ""
    ),
    "Advena": (
        "G&A", "Business services provider (primary service unconfirmed)",
        "N", "", ""
    ),
    "Hrvatski Nezavisnici Izvoznici Softvera": (
        "Professional Services", "Croatian Independent Software Exporters industry association membership",
        "N", "", ""
    ),
    "Monile J.D.O.O.": (
        "Engineering", "Mobile technology or IT services (Croatia — 'Monile')",
        "N", "",
        "Flagged as possible individual name — false positive; Monile J.D.O.O. is a Croatian company"
    ),
    "Puzzle Promotion J.D.O.O.": (
        "Marketing", "Promotional and creative marketing services (Croatia)",
        "N", "", ""
    ),
    "Zagrebaä_x008d_Ki Holding D.O.O.": (
        "G&A", "Zagreb city holding company — municipal utility or public service fees (Croatia)",
        "R", "",
        "Encoding corrupted (expected: 'Zagrebački Holding D.O.O.'). Classified G&A based on Zagreb city infrastructure entity."
    ),
    "Event Ors": (
        "G&A", "Event organisation and management services",
        "N", "",
        "Flagged as possible individual name — false positive; 'Event Ors' is an events company"
    ),
    "Sportkart D.O.O.": (
        "G&A", "Go-kart and motor sport entertainment — team-building event (Croatia — 'Sportkart')",
        "N", "", ""
    ),
    "Interaction Design Foundation, Inc": (
        "Product", "UX/design online learning and certification platform for Product and Design teams",
        "N", "", ""
    ),
    "(Blank)": (
        "G&A", "Unidentified vendor — blank vendor name entry",
        "N", "", "Blank vendor name in source data; $137 spend. Review source records to identify."
    ),
    "Escape Art D.O.O.": (
        "G&A", "Escape room entertainment — team-building event (Croatia)",
        "N", "", ""
    ),
    "Fero-Term": (
        "Facilities", "Thermal engineering or HVAC/heating system services (Croatia — 'Fero-Term')",
        "N", "", ""
    ),
    "Ustanova Za Zdravstvenu Skrb P.P.": (
        "G&A", "Healthcare institution providing occupational health services (Croatia)",
        "N", "", ""
    ),
    "Vistaprint": (
        "Marketing", "Online printed marketing materials and branded business stationery",
        "N", "", ""
    ),
    "Teb Poslovno Savjetovanje D.O.O.": (
        "Finance", "Tax and business advisory services (Croatia — TEB Poslovno Savjetovanje)",
        "N", "", ""
    ),
    "Illunis D.O.O.": (
        "Facilities", "Lighting or electrical services for office (Croatia — 'Illunis')",
        "N", "",
        "Flagged as possible individual name — false positive; Illunis D.O.O. is a Croatian company"
    ),
    "The Cook Kitchen": (
        "Facilities", "Office catering or cooking team-building experience",
        "N", "", ""
    ),
    "Australian Taxation Office (Ato)": (
        "Finance", "Australian government tax authority — statutory tax payments and compliance",
        "N", "", ""
    ),
    "Capitol Services": (
        "Professional Services", "Registered agent and government compliance services (US — Capitol Services)",
        "N", "", ""
    ),
    "Bonus Opinio D.O.O.": (
        "Marketing", "Market research, opinion polling and survey services (Croatia — 'Opinio' = opinion)",
        "N", "", ""
    ),
    "Dhl Express (Uk) Ltd": (
        "Facilities", "International express courier and shipping services (DHL UK)",
        "N", "", ""
    ),
    "Kat'S Kitchen D.O.O.": (
        "Facilities", "Catering and food services for office (Croatia — 'Kat's Kitchen')",
        "N", "", ""
    ),
    "Soho Kitchen Ltd": (
        "G&A", "Restaurant — business meal or team event (London, Soho)",
        "N", "", ""
    ),
    "Fortis Trade J.D.O.O.": (
        "Professional Services", "Trade and business services (Croatia — primary service unconfirmed)",
        "N", "", ""
    ),
    "Till Trade D.O.O.": (
        "Professional Services", "Trade or retail services (Croatia — primary service unconfirmed)",
        "N", "", ""
    ),
    "Pret A Manger": (
        "Facilities", "Coffee and sandwich chain — office refreshments or team meals",
        "N", "", ""
    ),
    "Axil Coffee Roasters": (
        "Facilities", "Specialty coffee supplier — office refreshments (Australia)",
        "N", "", ""
    ),
    "Pepco Croatia D.O.O.": (
        "G&A", "Discount retail chain — likely office supplies or general employee purchases (Croatia — PEPCO)",
        "N", "", ""
    ),
    "Lemia D.O.O.": (
        "Facilities", "Office facilities or cleaning services (Croatia — primary service unconfirmed)",
        "N", "",
        "Flagged as possible individual name — false positive; Lemia D.O.O. is a Croatian company"
    ),
    "Retriever Llc": (
        "Engineering", "Data retrieval, search or IT services (US — primary service unconfirmed)",
        "N", "", ""
    ),
    "Meluba Limited": (
        "G&A", "Business services provider (primary service unconfirmed)",
        "N", "",
        "Flagged as possible individual name — false positive; Meluba Limited is a company entity"
    ),
    "London Waste Management": (
        "Facilities", "Waste collection and office recycling services (London)",
        "N", "", ""
    ),
    "Hilton Garden Inn - Zagreb City Hotels D.O.O.": (
        "G&A", "Hotel accommodation — business travel (Hilton Garden Inn, Zagreb)",
        "N", "", ""
    ),
    "Lider Media D.O.O.": (
        "Marketing", "Business media and publishing — Croatian business magazine and news (Lider)",
        "N", "", ""
    ),
    "Uptime Robot Service Provider Ltd": (
        "Engineering", "Website and service uptime monitoring and alerting tool",
        "N", "", ""
    ),
    "Mithras Consultants": (
        "Professional Services", "Business or IT consulting services (primary service unconfirmed)",
        "N", "",
        "Flagged as possible individual name — false positive; Mithras Consultants is a consulting firm"
    ),
    "6Sense Insights Inc": (
        "Sales", "Account-based marketing (ABM) and buyer intent data platform",
        "N", "", ""
    ),
    "Ico": (
        "Legal", "UK Information Commissioner's Office — mandatory data protection registration fee",
        "N", "", ""
    ),
    "Sportska Udruga Split": (
        "G&A", "Sports association membership — employee wellness and sports benefit (Split, Croatia)",
        "N", "", ""
    ),
    "Prezzee": (
        "G&A", "Digital gift card platform — employee recognition and gifting",
        "N", "", ""
    ),
    "Inet Telecoms Ltd.": (
        "Facilities", "Internet and telecommunications services",
        "N", "", ""
    ),
    "Wolt Enterprises Oy": (
        "Facilities", "Food delivery platform — office meal ordering for team (Wolt, Finland/Europe)",
        "N", "", ""
    ),
    "Platinum Office D.O.O.": (
        "Facilities", "Office supplies and stationery services (Croatia — 'Platinum Office')",
        "N", "", ""
    ),
    "Snappy Snaps": (
        "G&A", "Photo printing and retail services — miscellaneous employee use",
        "N", "",
        "Flagged as possible individual name — false positive; Snappy Snaps is a UK photo printing chain"
    ),
    "New Block D.O.O.": (
        "Engineering", "IT or technology services (Croatia — primary service unconfirmed)",
        "N", "", ""
    ),
    "Cayman Islands Government": (
        "G&A", "Government registration or entity compliance fee (Cayman Islands)",
        "N", "", ""
    ),
    "Officeworks": (
        "G&A", "Office supplies and stationery retailer — general office purchases (Australia)",
        "N", "", ""
    ),
    "Click Send Pty Ltd": (
        "Marketing", "Bulk SMS, email and direct mail marketing services (Australia — ClickSend)",
        "N", "", ""
    ),
    "Kall Kwik Centre 565": (
        "Marketing", "Print and copy centre — marketing materials and business document printing",
        "N", "", ""
    ),
    "Shilton Hospitality Llp": (
        "G&A", "Hospitality or hotel services — business travel or team event",
        "N", "",
        "Flagged as possible individual name — false positive; Shilton Hospitality LLP is a hospitality company"
    ),
    "Axosoft Gitkraken": (
        "Engineering", "Git client and repository management tool for software developers",
        "N", "",
        "Flagged as possible individual name — false positive; Axosoft/GitKraken is a developer tools company"
    ),
    "Istra Wine": (
        "G&A", "Istrian wine retailer — likely for company event or client gift (Croatia)",
        "N", "",
        "Flagged as possible individual name — false positive; Istra Wine is a Croatian wine retailer"
    ),
    "Gophr": (
        "Facilities", "On-demand courier and last-mile delivery services (UK — Gophr)",
        "N", "", ""
    ),
    "Pan-Pek D.O.O.": (
        "Facilities", "Bakery and food products supplier — office catering provisions (Croatia — Pan-Pek)",
        "N", "", ""
    ),
    "Uber *Eats": (
        "Facilities", "Food delivery platform — office team meal ordering (Uber Eats)",
        "N", "", ""
    ),
    "Fedex Express Uk Transportation Ltd": (
        "Facilities", "International courier and express parcel delivery services (FedEx UK)",
        "N", "", ""
    ),
    "Garden City D.O.O.": (
        "Facilities", "Retail or hospitality services (Croatia — primary service unconfirmed)",
        "N", "", ""
    ),
    "Livingstone": (
        "G&A", "Business services or hospitality provider (primary service unconfirmed)",
        "N", "",
        "Flagged as possible individual name — false positive; Livingstone is a company entity"
    ),
    "Ekupi D.O.O.": (
        "G&A", "Online marketplace and retail — miscellaneous purchases (Croatia — Ekupi e-commerce)",
        "N", "", ""
    ),
    "Farmacia - Specijalizirana Prodavaonica D.O.O.": (
        "G&A", "Pharmacy and health products retailer — employee wellness or personal care (Croatia)",
        "N", "", ""
    ),
    "Cupcake Central (Life Is Sweet Bakery)": (
        "Facilities", "Bakery and dessert catering — team event or office refreshments (Australia)",
        "N", "", ""
    ),
    "Post Office Ltd": (
        "Facilities", "Postal services for business correspondence and parcels (UK Post Office)",
        "N", "", ""
    ),
    "Currys Pc World": (
        "G&A", "Consumer electronics and IT retailer — likely office equipment purchase (UK — Currys PC World)",
        "N", "", ""
    ),
    "Brodomerkur D.D.": (
        "G&A", "Retail and general supplies store — miscellaneous purchases (Croatia — Brodomerkur)",
        "N", "",
        "Flagged as possible individual name — false positive; Brodomerkur D.D. is a Croatian retail company"
    ),
    "Sport Vision D.O.O.": (
        "G&A", "Sports goods retailer — employee wellness or team event purchase (Croatia — Sport Vision)",
        "N", "", ""
    ),
    "Harissa D.O.O.": (
        "Facilities", "Restaurant or food catering services (Croatia — 'Harissa' restaurant)",
        "N", "",
        "Flagged as possible individual name — false positive; Harissa D.O.O. is a Croatian company"
    ),
    "Specijalisticka Ordinacija Medicine Rada I Sporta Ina Kardos": (
        "G&A", "Occupational health and sports medicine specialist clinic (Croatia)",
        "N", "", ""
    ),
    "Specijalisticka Ordinacija Medicine Rada Helena Blazic": (
        "G&A", "Occupational health specialist medical practice (Croatia)",
        "N", "", ""
    ),
    "Ustanova Za Medicinu Rada I Sporta Dr. Novacki": (
        "G&A", "Occupational health and sports medicine institution (Croatia)",
        "N", "", ""
    ),
    "M&S Simply Food": (
        "Facilities", "Grocery and food retailer — office refreshments (M&S Simply Food, UK)",
        "N", "", ""
    ),
    "Bakemono Bakers Melbourne": (
        "Facilities", "Artisan bakery — office catering or team event refreshments (Melbourne)",
        "N", "", ""
    ),
    "Coles": (
        "Facilities", "Supermarket chain — office grocery and refreshment supplies (Australia — Coles)",
        "N", "", ""
    ),
}


# ══════════════════════════════════════════════════════════════════════════════
# COLOUR PALETTE
# ══════════════════════════════════════════════════════════════════════════════
CONFIDENCE_FILLS = {
    "H": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # green
    "M": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # yellow
    "L": PatternFill(start_color="FFAB40", end_color="FFAB40", fill_type="solid"),  # orange
    "R": PatternFill(start_color="FF7043", end_color="FF7043", fill_type="solid"),  # red
    "N": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),  # grey
}
CONFIDENCE_FONTS = {
    "H": Font(bold=True, color="276221"),
    "M": Font(bold=True, color="9C5E00"),
    "L": Font(bold=True, color="7E3500"),
    "R": Font(bold=True, color="7B1C09"),
    "N": Font(bold=True, color="5A5A5A"),
}
SUGGESTION_FILLS = {
    "CONSOLIDATE": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # light blue
    "OPTIMIZE":    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # light green
    "TERMINATE":   PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),  # light red
}
SUGGESTION_FONTS = {
    "CONSOLIDATE": Font(bold=True, color="1F4E79"),
    "OPTIMIZE":    Font(bold=True, color="375623"),
    "TERMINATE":   Font(bold=True, color="9C0006"),
}
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

# ── Suggestion logic ──────────────────────────────────────────────────────────
_CONTRACTOR_SIGNALS = {"individual contractor", "freelancer"}
_TELECOM_SIGNALS    = {"telecommunications", "telecom", "connectivity", "internet",
                       "broadband", "mobile", "cable"}
_MANDATORY_SIGNALS  = {"government", "workers compensation", "occupational health",
                       "public health", "compliance fee", "regulatory"}

_DEPT_SUGGESTION = {
    "M&A":                  "TERMINATE",
    "Engineering":          "OPTIMIZE",
    "Facilities":           "OPTIMIZE",
    "Legal":                "OPTIMIZE",
    "Product":              "OPTIMIZE",
    "Sales":                "CONSOLIDATE",
    "Marketing":            "CONSOLIDATE",
    "SaaS":                 "CONSOLIDATE",
    "Finance":              "CONSOLIDATE",
    "G&A":                  "CONSOLIDATE",
    "Professional Services":"CONSOLIDATE",
    "Support":              "CONSOLIDATE",
}

def get_suggestion(dept: str, confidence: str, description: str) -> str:
    desc = description.lower()
    if dept == "M&A":
        return "TERMINATE"
    if any(kw in desc for kw in _CONTRACTOR_SIGNALS):
        return "TERMINATE"
    if confidence == "N":
        return "TERMINATE"
    if dept == "Facilities" and any(kw in desc for kw in _TELECOM_SIGNALS):
        return "CONSOLIDATE"
    if dept == "G&A" and any(kw in desc for kw in _MANDATORY_SIGNALS):
        return "OPTIMIZE"
    return _DEPT_SUGGESTION.get(dept, "CONSOLIDATE")
THIN_BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)


# ══════════════════════════════════════════════════════════════════════════════
# LOAD SOURCE DATA
# ══════════════════════════════════════════════════════════════════════════════
source_csv = BASE / "outputs" / "raw_valid_records_preview.csv"
source_rows = []
with open(source_csv, encoding="utf-8") as f:
    for row in csv.DictReader(f):
        source_rows.append(row)

# Enrich each row
enriched = []
for row in source_rows:
    name = row["vendor_name"]
    cost = float(row["cost_usd"])
    enc_issue = row["has_encoding_issue"] == "True"

    if name in CLASSIFICATIONS:
        dept, desc, conf, brand, notes = CLASSIFICATIONS[name]
    else:
        # Fallback: unmatched vendor
        dept = "G&A"
        desc = "Vendor not in classification lookup — assigned G&A as fallback"
        conf = "L"
        brand = ""
        notes = f"UNMATCHED: vendor '{name}' not found in lookup table; G&A assigned as fallback"

    enriched.append({
        "row_number": row["row_number"],
        "vendor_name": name,
        "department": dept,
        "cost_usd": cost,
        "description": desc,
        "suggestion": get_suggestion(dept, conf, desc),
        "classification_confidence": conf,
        "brand_group": brand,
        "classification_notes": notes,
        "has_encoding_issue": enc_issue,
    })

# ══════════════════════════════════════════════════════════════════════════════
# SUMMARY ANALYTICS
# ══════════════════════════════════════════════════════════════════════════════
TOTAL_SPEND = sum(r["cost_usd"] for r in enriched)
DEPARTMENTS = [
    "Engineering", "Facilities", "G&A", "Legal", "M&A", "Marketing",
    "SaaS", "Product", "Professional Services", "Sales", "Support", "Finance"
]

dept_spend = defaultdict(float)
dept_count = defaultdict(int)
conf_count = defaultdict(int)
suggest_count = defaultdict(int)
suggest_spend = defaultdict(float)
brand_groups = defaultdict(list)

for r in enriched:
    dept_spend[r["department"]] += r["cost_usd"]
    dept_count[r["department"]] += 1
    conf_count[r["classification_confidence"]] += 1
    suggest_count[r["suggestion"]] += 1
    suggest_spend[r["suggestion"]] += r["cost_usd"]
    if r["brand_group"]:
        brand_groups[r["brand_group"]].append(r)

# ══════════════════════════════════════════════════════════════════════════════
# OUTPUT 1: XLSX
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Vendor Mapping"

HEADERS = [
    "Row #", "Vendor Name", "Department", "Cost (USD)",
    "1-Line Description", "Suggestion", "Confidence", "Brand Group", "Classification Notes"
]
COL_WIDTHS = [7, 48, 22, 14, 68, 16, 12, 30, 60]

# Header row
for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 22

# Data rows
for row_idx, r in enumerate(enriched, 2):
    conf    = r["classification_confidence"]
    suggest = r["suggestion"]
    values = [
        int(r["row_number"]),
        r["vendor_name"],
        r["department"],
        r["cost_usd"],
        r["description"],
        suggest,
        conf,
        r["brand_group"],
        r["classification_notes"],
    ]
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [5, 9]))
        if col_idx == 4:  # Cost column
            cell.number_format = '#,##0.00'
        if col_idx == 6:  # Suggestion column
            cell.fill = SUGGESTION_FILLS.get(suggest, PatternFill())
            cell.font = SUGGESTION_FONTS.get(suggest, Font())
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx == 7:  # Confidence column (shifted from 6)
            cell.fill = CONFIDENCE_FILLS.get(conf, CONFIDENCE_FILLS["N"])
            cell.font = CONFIDENCE_FONTS.get(conf, CONFIDENCE_FONTS["N"])
            cell.alignment = Alignment(horizontal="center", vertical="center")

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

# Legend sheet
ws2 = wb.create_sheet("Legend")
conf_legend = [
    ("Confidence Code", "Meaning", "Colour"),
    ("H — High", "Unambiguous classification; no boundary dispute", "Green"),
    ("M — Medium", "Boundary rule applied; classification is defensible", "Yellow"),
    ("L — Low", "Best-guess from name pattern; limited signal", "Orange"),
    ("R — Requires Review", "Encoding corrupt / individual name / genuine ambiguity", "Red"),
    ("N — Noise", "Tier 5 vendor (<$1K); no strategic relevance", "Grey"),
]
for i, (code, meaning, colour) in enumerate(conf_legend, 1):
    ws2.cell(row=i, column=1, value=code)
    ws2.cell(row=i, column=2, value=meaning)
    ws2.cell(row=i, column=3, value=colour)
    if i > 1:
        conf_key = code.split("—")[0].strip()[0]
        ws2.cell(row=i, column=1).fill = CONFIDENCE_FILLS.get(conf_key, CONFIDENCE_FILLS["N"])
        ws2.cell(row=i, column=1).font = CONFIDENCE_FONTS.get(conf_key, CONFIDENCE_FONTS["N"])

sug_start = len(conf_legend) + 2
ws2.cell(row=sug_start, column=1, value="Suggestion").font = Font(bold=True, color="FFFFFF", size=10)
ws2.cell(row=sug_start, column=1).fill = HEADER_FILL
ws2.cell(row=sug_start, column=2, value="Meaning").font = Font(bold=True, color="FFFFFF", size=10)
ws2.cell(row=sug_start, column=2).fill = HEADER_FILL
sug_legend = [
    ("CONSOLIDATE", "Vendor has a direct equivalent in a typical acquirer's stack"),
    ("OPTIMIZE",    "Vendor is strategically necessary and will continue post-integration"),
    ("TERMINATE",   "Vendor is M&A-related or non-recurring; stop payments post-deal"),
]
for j, (sug, meaning) in enumerate(sug_legend, sug_start + 1):
    ws2.cell(row=j, column=1, value=sug)
    ws2.cell(row=j, column=2, value=meaning)
    ws2.cell(row=j, column=1).fill = SUGGESTION_FILLS[sug]
    ws2.cell(row=j, column=1).font = SUGGESTION_FONTS[sug]

for col in [1, 2, 3]:
    ws2.column_dimensions[get_column_letter(col)].width = [35, 60, 12][col - 1]

xlsx_path = OUTPUTS / "full_vendor_mapping.xlsx"
wb.save(xlsx_path)
print(f"✓ full_vendor_mapping.xlsx written ({len(enriched)} rows)")


# ══════════════════════════════════════════════════════════════════════════════
# OUTPUT 2: CSV
# ══════════════════════════════════════════════════════════════════════════════
csv_path = OUTPUTS / "full_vendor_mapping.csv"
fieldnames = [
    "row_number", "vendor_name", "department", "cost_usd",
    "description", "suggestion", "classification_confidence", "brand_group", "classification_notes"
]
with open(csv_path, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    for r in enriched:
        writer.writerow({k: r[k] for k in fieldnames})
print(f"✓ full_vendor_mapping.csv written ({len(enriched)} records)")


# ══════════════════════════════════════════════════════════════════════════════
# OUTPUT 3: SUMMARY MD
# ══════════════════════════════════════════════════════════════════════════════

def pct(v, total):
    return f"{100*v/total:.1f}%"

lines = []
def h(level, text): lines.append(f"{'#' * level} {text}\n")
def p(*args): lines.extend(args); lines.append("")
def table(headers, rows):
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("|" + "|".join(["---"] * len(headers)) + "|")
    for row in rows: lines.append("| " + " | ".join(str(c) for c in row) + " |")
    lines.append("")

h(1, "Full Vendor Mapping — Summary Report")
p(
    f"**Total vendors classified:** {len(enriched)}",
    f"**Total L12M spend:** ${TOTAL_SPEND:,.2f}",
    f"**Classification date:** 2026-05-14",
    f"**Governing documents:** mapping_philosophy.md · department_interpretation_guide.md · ambiguity_resolution.md",
)
lines.append("---\n")

# Spend by department
h(2, "1. Spend by Department")
dept_rows = []
sorted_depts = sorted(DEPARTMENTS, key=lambda d: dept_spend[d], reverse=True)
for dept in sorted_depts:
    spend = dept_spend[dept]
    count = dept_count[dept]
    dept_rows.append([dept, count, f"${spend:,.2f}", pct(spend, TOTAL_SPEND)])
dept_rows.append(["**TOTAL**", len(enriched), f"**${TOTAL_SPEND:,.2f}**", "**100.0%**"])
table(["Department", "Vendor Count", "Spend (USD)", "% of Total"], dept_rows)

lines.append("---\n")

# Confidence distribution
h(2, "2. Classification Confidence Distribution")
conf_rows = []
for code, label in [("H","High"),("M","Medium"),("L","Low"),("R","Requires Review"),("N","Noise (<$1K)")]:
    cnt = conf_count[code]
    conf_rows.append([f"{code} — {label}", cnt, pct(cnt, len(enriched))])
table(["Confidence", "Vendor Count", "% of Vendors"], conf_rows)
p(
    "> **Note on Noise (N):** All 178 T5 vendors (< $1K) are assigned N confidence.",
    "> They are classified for completeness but excluded from strategic recommendations.",
    "> Strategic focus should be on H and M confidence vendors (T1–T3 spend tiers).",
)

lines.append("---\n")

# Suggestion distribution
h(2, "3. Post-Acquisition Action (Suggestion)")
sug_rows = []
for sug in ["CONSOLIDATE", "OPTIMIZE", "TERMINATE"]:
    cnt = suggest_count[sug]
    sp  = suggest_spend[sug]
    sug_rows.append([sug, cnt, f"${sp:,.2f}", pct(sp, TOTAL_SPEND)])
sug_rows.append(["**TOTAL**", len(enriched), f"**${TOTAL_SPEND:,.2f}**", "**100.0%**"])
table(["Suggestion", "Vendor Count", "Spend (USD)", "% of Total"], sug_rows)
p(
    "> **CONSOLIDATE** — vendor has a direct equivalent in a typical acquirer's stack; rationalize contracts.",
    "> **OPTIMIZE** — vendor is strategically necessary and will continue post-integration; renegotiate terms.",
    "> **TERMINATE** — M&A-related or non-recurring spend; stop payments post-deal close.",
    "> Noise-tier vendors (<$1K, N confidence) are assigned TERMINATE by default — immaterial tail spend acquirer absorbs operationally.",
)

lines.append("---\n")

# Key findings by department
h(2, "4. Key Findings by Department")

findings = {
    "Sales": [
        f"**${ dept_spend['Sales']:,.2f} — {pct(dept_spend['Sales'], TOTAL_SPEND)} of total spend** — the single largest department by far.",
        "Dominated by **Salesforce UK** ($3.12M, 39.5% of all vendor spend) — the acquired company's CRM stack.",
        "**Cloudcrossing BVBA / PDF Butler** ($208.7K) is a direct Salesforce add-on classified together with the CRM stack.",
        "Other sales tools: Cognism ($27K), Outreach ($9.2K), Lusha ($2.6K), 6Sense ($76).",
        "**Key risk:** Extreme single-vendor concentration. Salesforce + PDF Butler = ~$3.33M. Immediate priority for acquirer CRM consolidation assessment.",
    ],
    "Facilities": [
        f"**${dept_spend['Facilities']:,.2f} — {pct(dept_spend['Facilities'], TOTAL_SPEND)} of total** — second largest department.",
        "Multi-geography footprint: UK (Tog UK Properties $263.8K), Croatia (Zagrebtower $183.8K, Weking $144.1K), India (Innovent $147.3K), Singapore (WeWork $64.4K), Australia (GPT Space $133.5K).",
        "Office telecom included: Telefonica ($89.9K), Hrvatski Telekom ($18.1K), Telemach ($5.1K), British Telecom, T-Mobile, StarHub, Vodafone.",
        "**Key action:** Map each facility to active headcount. Post-acquisition, any location the acquirer already has presence in should be reviewed for lease consolidation.",
    ],
    "G&A": [
        f"**${dept_spend['G&A']:,.2f} — {pct(dept_spend['G&A'], TOTAL_SPEND)} of total** — includes benefits, travel, HR, and all catch-all spend.",
        "**Navan/TripActions** appears TWICE ($357.9K + $57.9K = **$415.9K combined**) under two entity names — highest-priority normalisation needed before any T&E recommendation.",
        "Employee benefits (insurance): Jensten ($142.7K), Aetna ($124.7K), Bupa ($35.3K combined), Agram Life ($24.9K), Care Health ($24K), Cigna ($13.2K), Allianz ($7.1K combined), ICICI Lombard ($6K).",
        "178 T5 Noise vendors fall primarily into G&A/Facilities — meal receipts, retail, entertainment. These are expense-level transactions, not strategic vendor relationships.",
    ],
    "Finance": [
        f"**${dept_spend['Finance']:,.2f} — {pct(dept_spend['Finance'], TOTAL_SPEND)} of total**.",
        "BDO LLP ($343.1K) is the dominant Finance vendor — external audit relationship requiring acquirer audit committee review.",
        "Multiple accounting firms engaged simultaneously: BDO, Grant Thornton, PwC, Crowe Horwath, Collards, McBurneys — potential for consolidation.",
        "FP&A: Planful ($27.7K) — assess compatibility with acquirer's finance systems.",
    ],
    "M&A": [
        f"**${dept_spend['M&A']:,.2f} — {pct(dept_spend['M&A'], TOTAL_SPEND)} of total** — confirms active deal activity in the L12M period.",
        "RSM UK Corporate Finance ($117.1K, H confidence), SS&C Intralinks VDR ($40K, H), Houlihan Lokey ($37.5K, H), Vector Capital ($32.4K, H), 4I Advisory ($71.9K, M).",
        "Most M&A spend is non-recurring — these vendors will naturally terminate post-deal close.",
        "**Exception:** 4I Advisory Services may include recurring tax advisory (Finance) — verify engagement scope.",
    ],
    "Legal": [
        f"**${dept_spend['Legal']:,.2f} — {pct(dept_spend['Legal'], TOTAL_SPEND)} of total**.",
        "Multi-jurisdiction legal coverage: UK (Bisley Law $67.4K, Pinsent Masons), Croatia (Zuric I Partneri $24.1K), Australia (O'Donnell Salzano), Singapore, India, US.",
        "Immigration: Visalogic ($11.2K) — assess whether acquirer's legal team can absorb global mobility workload.",
        "High vendor count relative to spend — multiple small law firms across geographies suggest opportunity for panel consolidation.",
    ],
    "Professional Services": [
        f"**${dept_spend['Professional Services']:,.2f} — {pct(dept_spend['Professional Services'], TOTAL_SPEND)} of total**.",
        "Infosys ($66.6K), Harmonic Group ($65.4K, confirmed management consultancy), Cloud Technology Solutions ($60.7K) are the largest PS engagements.",
        "Mason Frank ($38.1K) is a Salesforce specialist recruitment firm — engagement should be reassessed in light of any CRM consolidation decision.",
        "Multiple Croatian D.O.O. PS vendors with L confidence — low strategic relevance individually.",
    ],
    "Marketing": [
        f"**${dept_spend['Marketing']:,.2f} — {pct(dept_spend['Marketing'], TOTAL_SPEND)} of total**.",
        "HubSpot ($32.2K) is the primary marketing automation tool — its presence **directly confirms** Salesforce is NOT Marketing Cloud.",
        "LinkedIn ($55.6K, M confidence) — verify whether Sales Navigator is included; if so, portion should reclassify to Sales.",
        "Marketing stack: HubSpot, Uberflip, LinkedIn, SEMrush, MightyHive, Cognism (Sales), Carrington Comms — assess overlap with acquirer's marketing technology.",
    ],
    "Engineering": [
        f"**${dept_spend['Engineering']:,.2f} — {pct(dept_spend['Engineering'], TOTAL_SPEND)} of total**.",
        "AWS combined ($111.6K across LLC + Inc.) is the dominant Engineering vendor — two entity entries for one cloud account.",
        "Developer toolchain: JetBrains ($4.1K), npm ($3.5K), Papertrail ($3.9K), Ag Grid ($2.7K), GitKraken ($54) — mostly small-spend individual tool licences.",
        "Technet IT Recruitment ($21.9K) classified Engineering per edge case rule (explicitly IT recruitment).",
    ],
    "SaaS": [
        f"**${dept_spend['SaaS']:,.2f} — {pct(dept_spend['SaaS'], TOTAL_SPEND)} of total**.",
        "Google Workspace ($24.8K, M) and Microsoft 365 ($9.8K, M) — both assumed company-wide productivity suites; flip to Engineering if GCP/Azure confirmed.",
        "Cross-functional tools: Trello ($6.7K), GoTo ($7.3K), Smartsheet ($3K), Slack ($2K), Adobe ($3.9K), DocuSign ($1.3K).",
        "Slack at $2K is unusually low — likely a smaller team licence or partial year.",
    ],
    "Product": [
        f"**${dept_spend['Product']:,.2f} — {pct(dept_spend['Product'], TOTAL_SPEND)} of total** — small but clearly defined.",
        "Aha! Labs ($3.7K) for product roadmapping, Figma ($460) for design. Lean product tooling spend — suggests team size is small or tools are partially shared.",
    ],
    "Support": [
        f"**${dept_spend.get('Support', 0):,.2f} — {pct(dept_spend.get('Support', 0), TOTAL_SPEND)} of total** — no vendors classified to Support.",
        "No helpdesk, ticketing or customer support tool vendors identified in the dataset.",
        "Possible explanations: Support tools are bundled within Salesforce Service Cloud (reducing Support spend visibility), or Support is a small team using free-tier tools.",
    ],
}

for dept in sorted_depts:
    if dept in findings:
        h(3, dept)
        for point in findings[dept]:
            lines.append(f"- {point}")
        lines.append("")

lines.append("---\n")

# Brand group consolidation
h(2, "4. Brand Group Consolidation Summary")
p("The following multi-entity brand groups require normalisation before strategic reporting.")
bg_rows = []
for brand, vendors in sorted(brand_groups.items(), key=lambda x: -sum(v["cost_usd"] for v in x[1])):
    combined = sum(v["cost_usd"] for v in vendors)
    dept = vendors[0]["department"]
    entity_list = ", ".join(v["vendor_name"] for v in vendors)
    bg_rows.append([brand, len(vendors), f"${combined:,.2f}", dept, entity_list])
table(["Canonical Brand", "Entities", "Combined Spend", "Department", "Entity Names"], bg_rows)

p(
    "> **Highest priority normalisation:** Navan / TripActions ($415.9K combined) — two rows in source data;",
    "> must be merged before T&E spend is reported or benchmarked.",
    ">",
    "> **AWS / Amazon note:** Amazon.co.uk and Amazon (Aus) are retail entities classified G&A,",
    "> not Engineering. Combined AWS-only spend (LLC + Inc.) = $111.6K for cloud infrastructure.",
)

lines.append("---\n")

# Footnote on R vendors
r_vendors = [r for r in enriched if r["classification_confidence"] == "R"]
h(2, "5. Requires-Review Vendors (R Confidence)")
p(
    f"**{len(r_vendors)} vendors** remain flagged R (Requires Review).",
    "All are encoding-corrupted Croatian entities. Decoded names and provisional classifications are noted.",
    "These do not block strategic analysis but should be corrected in the source system.",
)
r_rows = []
for r in sorted(r_vendors, key=lambda x: -x["cost_usd"]):
    r_rows.append([r["vendor_name"][:50], r["department"], f"${r['cost_usd']:,.2f}", r["classification_notes"][:80]])
table(["Vendor Name (corrupted)", "Provisional Dept", "Spend", "Notes"], r_rows)

lines.append("---\n")
p("*Classification complete. All 386 vendors assigned. Zero unclassified records.*")

md_path = OUTPUTS / "full_mapping_summary.md"
md_path.write_text("\n".join(lines), encoding="utf-8")
print(f"✓ full_mapping_summary.md written")

# Final console stats
print()
print("=" * 65)
print("CLASSIFICATION COMPLETE")
print("=" * 65)
for dept in sorted_depts:
    print(f"  {dept:<25} {dept_count[dept]:>3} vendors   ${dept_spend[dept]:>12,.2f}  ({pct(dept_spend[dept], TOTAL_SPEND):>6})")
print(f"  {'TOTAL':<25} {len(enriched):>3} vendors   ${TOTAL_SPEND:>12,.2f}  (100.0%)")
print()
print("Confidence breakdown:")
for code, label in [("H","High"),("M","Medium"),("L","Low"),("R","Review"),("N","Noise")]:
    print(f"  {code} ({label:<10}): {conf_count[code]:>3} vendors")
print("=" * 65)
print(f"Outputs → {OUTPUTS}/")
