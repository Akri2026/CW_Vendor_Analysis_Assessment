#!/usr/bin/env python3
"""
fix_xlsx_suggestions.py
Overwrites the Suggestion column F in full_vendor_mapping.xlsx with the
correct values from full_vendor_mapping.csv (which is the authoritative source).

The previous add_suggestions.py pass wrote wrong values for 174 noise-tier rows
because openpyxl's iter_rows(values_only=True) returned None for those cells'
confidence column, causing the department default to fire instead of the "N" → TERMINATE rule.
"""

import csv
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE    = Path(__file__).resolve().parent.parent
OUTPUTS = BASE / "outputs" / "mappings"

# ── Styles ────────────────────────────────────────────────────────────────────
SUGGESTION_FILLS = {
    "CONSOLIDATE": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "OPTIMIZE":    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "TERMINATE":   PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
}
SUGGESTION_FONTS = {
    "CONSOLIDATE": Font(bold=True, color="1F4E79"),
    "OPTIMIZE":    Font(bold=True, color="375623"),
    "TERMINATE":   Font(bold=True, color="9C0006"),
}
THIN_BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

# ── Read correct suggestions from CSV ─────────────────────────────────────────
csv_path = OUTPUTS / "full_vendor_mapping.csv"
csv_suggestions = []
with open(csv_path, encoding="utf-8") as f:
    for row in csv.DictReader(f):
        csv_suggestions.append(row["suggestion"])

print(f"Read {len(csv_suggestions)} suggestions from CSV")

# ── Patch XLSX ────────────────────────────────────────────────────────────────
xlsx_path = OUTPUTS / "full_vendor_mapping.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
ws = wb["Vendor Mapping"]

# Verify column F header is "Suggestion"
col_f_header = ws.cell(row=1, column=6).value
assert col_f_header and "suggestion" in col_f_header.lower(), f"Unexpected header in F1: {col_f_header!r}"

# Overwrite each data row's column F with the correct suggestion from CSV
for i, suggestion in enumerate(csv_suggestions):
    row_idx = i + 2  # row 1 is header; data starts at row 2
    cell = ws.cell(row=row_idx, column=6, value=suggestion)
    cell.fill      = SUGGESTION_FILLS.get(suggestion, PatternFill())
    cell.font      = SUGGESTION_FONTS.get(suggestion, Font())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = THIN_BORDER

wb.save(xlsx_path)
print(f"✓ full_vendor_mapping.xlsx — column F overwritten with correct suggestions ({len(csv_suggestions)} rows)")

# ── Verify distribution ───────────────────────────────────────────────────────
dist = Counter(csv_suggestions)
total = len(csv_suggestions)
print("\nCorrected suggestion distribution (from CSV → applied to XLSX):")
print(f"  {'Suggestion':<14} {'Vendors':>7}   {'% of Vendors':>13}")
print(f"  {'-'*40}")
for sug in ["CONSOLIDATE", "OPTIMIZE", "TERMINATE"]:
    pct = 100 * dist[sug] / total
    print(f"  {sug:<14} {dist[sug]:>7}   {pct:>12.1f}%")
print(f"  {'-'*40}")
print(f"  {'TOTAL':<14} {total:>7}   {'100.0%':>13}")
