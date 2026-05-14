#!/usr/bin/env python3
"""
add_suggestions.py
Patches the existing full_vendor_mapping.xlsx and full_vendor_mapping.csv
by inserting a Suggestion column (CONSOLIDATE / OPTIMIZE / TERMINATE) at
column F, without touching the classification columns.

Governing logic (priority order):
  1. department == "M&A"                              → TERMINATE
  2. description contains individual contractor signal → TERMINATE
  3. confidence == "N" (noise, <$1K)                  → TERMINATE
  4. Facilities + telecom signal in description        → CONSOLIDATE
  5. G&A + mandatory/regulatory signal in description  → OPTIMIZE
  6. Department default:
       OPTIMIZE  → Engineering, Facilities, Legal, Product
       CONSOLIDATE → Sales, Marketing, SaaS, Finance, G&A, Professional Services, Support
"""

import csv
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE    = Path(__file__).resolve().parent.parent
OUTPUTS = BASE / "outputs" / "mappings"

# ── Suggestion logic ──────────────────────────────────────────────────────────
_CONTRACTOR_SIGNALS = {"individual contractor", "freelancer"}
_TELECOM_SIGNALS    = {
    "telecommunications", "telecom", "connectivity",
    "internet", "broadband", "mobile", "cable",
}
_MANDATORY_SIGNALS  = {
    "government", "workers compensation", "occupational health",
    "public health", "compliance fee", "regulatory",
}

_DEPT_SUGGESTION = {
    "M&A":                   "TERMINATE",
    "Engineering":           "OPTIMIZE",
    "Facilities":            "OPTIMIZE",
    "Legal":                 "OPTIMIZE",
    "Product":               "OPTIMIZE",
    "Sales":                 "CONSOLIDATE",
    "Marketing":             "CONSOLIDATE",
    "SaaS":                  "CONSOLIDATE",
    "Finance":               "CONSOLIDATE",
    "G&A":                   "CONSOLIDATE",
    "Professional Services": "CONSOLIDATE",
    "Support":               "CONSOLIDATE",
}


def get_suggestion(dept: str, confidence: str, description: str) -> str:
    desc = description.lower()
    if dept == "M&A":
        return "TERMINATE"
    if any(kw in desc for kw in _CONTRACTOR_SIGNALS):
        return "TERMINATE"
    if confidence == "N":
        return "TERMINATE"
    if dept == "Facilities" and any(kw in desc for kw in _TELECOM_SIGNALS):
        return "CONSOLIDATE"
    if dept == "G&A" and any(kw in desc for kw in _MANDATORY_SIGNALS):
        return "OPTIMIZE"
    return _DEPT_SUGGESTION.get(dept, "CONSOLIDATE")


# ── Styles ────────────────────────────────────────────────────────────────────
SUGGESTION_FILLS = {
    "CONSOLIDATE": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "OPTIMIZE":    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "TERMINATE":   PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
}
SUGGESTION_FONTS = {
    "CONSOLIDATE": Font(bold=True, color="1F4E79"),
    "OPTIMIZE":    Font(bold=True, color="375623"),
    "TERMINATE":   Font(bold=True, color="9C0006"),
}
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

# ── Patch XLSX ────────────────────────────────────────────────────────────────
xlsx_path = OUTPUTS / "full_vendor_mapping.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
ws = wb["Vendor Mapping"]

# Current layout (before any changes):
#   A=Row#  B=Vendor Name  C=Department  D=Cost  E=Description  F=Confidence  G=Brand  H=Notes
# Read dept/conf/desc from existing columns BEFORE inserting anything.
row_suggestions = []
for row in ws.iter_rows(min_row=2, values_only=True):
    dept = str(row[2] or "")   # col C, index 2
    desc = str(row[4] or "")   # col E, index 4
    conf = str(row[5] or "")   # col F, index 5
    row_suggestions.append(get_suggestion(dept, conf, desc))

# Insert blank column at position 6 (→ existing F becomes G, G→H, H→I)
ws.insert_cols(6)
ws.column_dimensions["F"].width = 16

# Header
hdr = ws.cell(row=1, column=6, value="Suggestion")
hdr.fill = HEADER_FILL
hdr.font = HEADER_FONT
hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
hdr.border = THIN_BORDER

# Data cells
for row_idx, suggestion in enumerate(row_suggestions, 2):
    cell = ws.cell(row=row_idx, column=6, value=suggestion)
    cell.fill = SUGGESTION_FILLS.get(suggestion, PatternFill())
    cell.font = SUGGESTION_FONTS.get(suggestion, Font())
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

# Extend autofilter to cover new column I
ws.auto_filter.ref = "A1:I1"

# ── Update Legend sheet ───────────────────────────────────────────────────────
ws2 = wb["Legend"]
sug_start = ws2.max_row + 2

for col_idx, (val, is_header) in enumerate([
    ("Suggestion", True), ("Meaning", True)
], 1):
    c = ws2.cell(row=sug_start, column=col_idx, value=val)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT

sug_legend = [
    ("CONSOLIDATE", "Vendor has a direct equivalent in a typical acquirer's stack; rationalize contracts"),
    ("OPTIMIZE",    "Vendor is strategically necessary and will continue post-integration; renegotiate terms"),
    ("TERMINATE",   "M&A-related or non-recurring spend; stop payments post-deal close"),
]
for j, (sug, meaning) in enumerate(sug_legend, sug_start + 1):
    c1 = ws2.cell(row=j, column=1, value=sug)
    c1.fill = SUGGESTION_FILLS[sug]
    c1.font = SUGGESTION_FONTS[sug]
    ws2.cell(row=j, column=2, value=meaning)

wb.save(xlsx_path)
print(f"✓ full_vendor_mapping.xlsx — Suggestion column inserted at column F ({len(row_suggestions)} rows)")

# ── Patch CSV ─────────────────────────────────────────────────────────────────
csv_path = OUTPUTS / "full_vendor_mapping.csv"
rows = []
with open(csv_path, encoding="utf-8") as f:
    reader = csv.DictReader(f)
    old_fieldnames = list(reader.fieldnames)
    for row in reader:
        rows.append(row)

for row in rows:
    dept = row.get("department", "")
    conf = row.get("classification_confidence", "")
    desc = row.get("description", "")
    row["suggestion"] = get_suggestion(dept, conf, desc)

# Insert "suggestion" after "description" in fieldnames
new_fieldnames = []
for fn in old_fieldnames:
    new_fieldnames.append(fn)
    if fn == "description":
        new_fieldnames.append("suggestion")

with open(csv_path, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=new_fieldnames)
    writer.writeheader()
    writer.writerows(rows)

print(f"✓ full_vendor_mapping.csv  — suggestion column added ({len(rows)} records)")

# ── Distribution summary ──────────────────────────────────────────────────────
dist  = Counter(r["suggestion"] for r in rows)
spend = {
    sug: sum(float(r["cost_usd"]) for r in rows if r["suggestion"] == sug)
    for sug in ["CONSOLIDATE", "OPTIMIZE", "TERMINATE"]
}
total = sum(float(r["cost_usd"]) for r in rows)

print("\nSuggestion distribution:")
print(f"  {'Suggestion':<14} {'Vendors':>7}   {'Spend (USD)':>14}   {'% of Total':>10}")
print(f"  {'-'*55}")
for sug in ["CONSOLIDATE", "OPTIMIZE", "TERMINATE"]:
    pct = 100 * spend[sug] / total
    print(f"  {sug:<14} {dist[sug]:>7}   ${spend[sug]:>13,.2f}   {pct:>9.1f}%")
print(f"  {'-'*55}")
print(f"  {'TOTAL':<14} {sum(dist.values()):>7}   ${total:>13,.2f}   {'100.0%':>10}")
